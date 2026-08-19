"""Production vocabulary labelling (the LLM route, post-gating GREEN LIGHT).

Labels the unmatched Plaid merchant vocabulary with the context-enriched
two-model pipeline and gates every string into a quality tier. Accepted
labels become T4 lookup data; nothing enters the dictionary silently.

Tiers (thresholds from measured calibration -- CLAUDE.md section 6):
    auto_accept        models agree AND sonnet confidence >= 0.9
    accepted           models agree AND sonnet confidence >= 0.7
    abstain_confirmed  both models say unclassified_other
    needs_review       everything else (disagreement / low confidence)

Usage:
    python src/production_labelling.py fetch [N]     # top-N unmatched strings + evidence
    python src/production_labelling.py label [haiku|sonnet]
    python src/production_labelling.py tiebreak       # Opus over the needs_review queue
    python src/production_labelling.py gate           # -> outputs/production_labels.csv + stats
    python src/production_labelling.py review-sheet   # workbook for risk-boundary strings
    python src/production_labelling.py apply-review [completed.xlsx]

Tranche runbook (run from the repo root, venv active, after `gcloud auth login`
if BigQuery credentials have expired):

    python src/production_labelling.py fetch 20000    # re-selects top N; already-labelled strings resume for free
    python src/production_labelling.py label haiku    # only labels new strings (resumable, checkpointed)
    python src/production_labelling.py label sonnet
    python src/production_labelling.py gate           # first gate: exposes the new needs_review queue
    python src/production_labelling.py tiebreak       # Opus labels ONLY current needs_review strings
    python src/production_labelling.py gate           # re-gate with the tiebreak applied
    # re-apply every archived human review, then the policy leaves only NEW risk-boundary strings:
    python src/production_labelling.py apply-review data/production_review_tranche1_completed.xlsx
    python src/production_labelling.py review-sheet   # workbook with only the new human work
    # ... human pass, save as outputs/production_review_completed.xlsx ...
    python src/production_labelling.py apply-review
    # then snapshot: cp outputs/production_labels.csv data/production_labels_trancheN.csv
    #                cp outputs/production_review_completed.xlsx data/production_review_trancheN_completed.xlsx
"""
import csv
import json
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from gating_experiment import (  # noqa: E402
    MODELS, ROOT, OUT_DIR, build_system_prompt, build_tool_schema,
    load_crosswalk, load_example_merchants,
)
from build_tail_eval import TAIL_ADDENDUM, POPULATION_QUERY, bq_json  # noqa: E402

STRINGS_CSV = OUT_DIR / "production_strings.csv"
EVIDENCE_JSON = OUT_DIR / "production_evidence.json"
PREDICTIONS = {k: OUT_DIR / f"production_predictions_{k}.csv" for k in MODELS}
OPUS_PREDICTIONS = OUT_DIR / "production_predictions_opus.csv"
LABELS_CSV = OUT_DIR / "production_labels.csv"
BATCH = 20
DEFAULT_N = 5000

# Tiebreaker for needs_review strings: a third, stronger model breaks 2-of-3
# majorities. Local config (not in gating MODELS -- the two-model experiments
# stay two-model). Opus 5: no sampling params; thinking suppressed under
# forced tool_choice, same as Sonnet 5.
TIEBREAK_CFG = {"id": "claude-opus-5", "max_tokens": 16000, "extra": {}}

# strings already gold-labelled by human adjudication are excluded here --
# their labels come from data/, not from this pipeline
GOLD_FILES = [ROOT / "data" / "gold_merchant_labels.csv", ROOT / "data" / "gold_tail_labels.csv"]


def fetch(n):
    print("Querying unmatched Plaid merchant population...", file=sys.stderr)
    pop = sorted(((r["m"], int(r["n"])) for r in bq_json(POPULATION_QUERY)), key=lambda x: -x[1])
    already_gold = set()
    for gf in GOLD_FILES:
        if gf.exists():
            already_gold |= {r["merchant"].strip().lower() for r in csv.DictReader(open(gf))}
    total_vol = sum(v for _, v in pop)
    chosen = [(m, v) for m, v in pop if m not in already_gold][:n]
    print(f"Selected top {len(chosen)} strings covering "
          f"{sum(v for _, v in chosen) / total_vol:.1%} of unmatched volume "
          f"({len(already_gold & {m for m, _ in pop})} gold-labelled strings excluded)", file=sys.stderr)

    with open(STRINGS_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["merchant", "plaid_n"])
        w.writerows(chosen)

    def q(s):
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    evidence = []
    CHUNK = 2000  # keep each IN-list query a sane size
    for i in range(0, len(chosen), CHUNK):
        part = chosen[i:i + CHUNK]
        in_list = ", ".join(q(m) for m, _ in part)
        print(f"Evidence chunk {i // CHUNK + 1}/{(len(chosen) + CHUNK - 1) // CHUNK}...", file=sys.stderr)
        evidence += bq_json(f"""
SELECT LOWER(TRIM(merchant_name)) AS m,
       ROUND(COUNTIF(amount < 0) / COUNT(*), 2) AS pct_credit,
       ROUND(APPROX_QUANTILES(ABS(amount), 2)[OFFSET(1)], 2) AS median_amount,
       APPROX_TOP_COUNT(credit_category_detailed, 2) AS cats,
       APPROX_TOP_COUNT(COALESCE(original_description, transaction_name), 3) AS descs
FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
  AND LOWER(TRIM(merchant_name)) IN ({in_list})
GROUP BY 1
""")
    EVIDENCE_JSON.write_text(json.dumps(evidence))
    print(f"Wrote {STRINGS_CSV} and {EVIDENCE_JSON}", file=sys.stderr)


def load_evidence():
    ev = {}
    for r in json.loads(EVIDENCE_JSON.read_text()):
        ev[r["m"]] = {
            "pct_credit": float(r["pct_credit"]),
            "median_amount": float(r["median_amount"]),
            "cats": " · ".join(f"{c['value']} {int(c['count'])}x" for c in r["cats"][:2]),
            "descs": " | ".join(f"\"{d['value'].strip()[:70]}\"" for d in r["descs"][:3]),
        }
    return ev


def run_labelling(cfg, rows, out_path):
    import anthropic

    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = build_system_prompt(leaves, gen_of, notes_of, load_example_merchants()) + TAIL_ADDENDUM
    tool = build_tool_schema(leaves)
    ev = load_evidence()

    # resumable: skip strings already predicted (rerun-safe after interruption)
    predictions = {}
    if out_path.exists():
        predictions = {r["merchant"]: {"leaf": r["llm_leaf"], "conf": r["llm_confidence"]}
                       for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)
    todo = [r for r in rows if r["merchant"] not in predictions]

    client = anthropic.Anthropic()
    n_batches = (len(todo) + BATCH - 1) // BATCH

    def render(i, r):
        e = ev.get(r["merchant"], {})
        return (f"{i}. merchant: {r['merchant']}\n"
                f"   plaid_native_categories: {e.get('cats', 'n/a')}\n"
                f"   pct_credit: {e.get('pct_credit', 'n/a')} | median_amount_gbp: {e.get('median_amount', 'n/a')}\n"
                f"   top_raw_narratives: {e.get('descs', 'n/a')}")

    def flush():
        # write the union of this run's rows and every previously-labelled
        # merchant -- a later tranche must never erase an earlier tranche's
        # predictions from the shared file
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["merchant", "llm_leaf", "llm_confidence"])
            w.writeheader()
            seen = set()
            for r in rows:
                seen.add(r["merchant"])
                p = predictions.get(r["merchant"])
                w.writerow({"merchant": r["merchant"],
                            "llm_leaf": p["leaf"] if p else "",
                            "llm_confidence": p["conf"] if p else ""})
            for m, p in predictions.items():
                if m not in seen:
                    w.writerow({"merchant": m, "llm_leaf": p["leaf"], "llm_confidence": p["conf"]})

    def classify_batch(batch, tag):
        user_msg = "Classify each of these merchant strings using the evidence provided:\n\n" + \
            "\n".join(render(j + 1, r) for j, r in enumerate(batch))
        response = client.messages.create(
            model=cfg["id"], max_tokens=cfg["max_tokens"],
            system=[{"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral", "ttl": "1h"}}],
            tools=[tool], tool_choice={"type": "tool", "name": "submit_classifications"},
            messages=[{"role": "user", "content": user_msg}], **cfg["extra"],
        )
        if response.stop_reason == "max_tokens":
            print(f"  WARNING: [{tag}] truncated", file=sys.stderr)
        tool_use = next((b for b in response.content if b.type == "tool_use"), None)
        if tool_use is None:
            return {}
        by_string = {r["merchant"].strip().lower(): r["merchant"] for r in batch}
        got = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            echoed = (res.get("merchant") or "").strip().lower()
            merchant = None
            if isinstance(idx, int) and 1 <= idx <= len(batch):
                candidate = batch[idx - 1]["merchant"]
                if candidate.strip().lower() == echoed or echoed not in by_string:
                    merchant = candidate
            if merchant is None and echoed in by_string:
                merchant = by_string[echoed]
            if merchant is not None:
                got[merchant] = {"leaf": res.get("detailed_category"), "conf": res.get("confidence")}
        return got

    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 10 == 1:
            print(f"[{cfg['id']}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:03d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if r["merchant"] not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:03d}_r{attempt}"))
        if num % 25 == 0:
            flush()  # checkpoint

    flush()
    missing = sum(1 for r in rows if r["merchant"] not in predictions)
    print(f"Wrote {out_path} ({len(predictions)} labelled, {missing} missing)", file=sys.stderr)


def label(model_key):
    rows = list(csv.DictReader(open(STRINGS_CSV)))
    run_labelling(MODELS[model_key], rows, PREDICTIONS[model_key])


def tiebreak():
    """Run the tiebreaker model over the needs_review strings only."""
    labels = list(csv.DictReader(open(LABELS_CSV)))
    rows = [{"merchant": r["merchant"], "plaid_n": r["plaid_n"]}
            for r in labels if r["tier"] == "needs_review"]
    print(f"Tiebreaking {len(rows)} needs_review strings with {TIEBREAK_CFG['id']}", file=sys.stderr)
    run_labelling(TIEBREAK_CFG, rows, OPUS_PREDICTIONS)


RISK_GENERALS = {"gambling", "high_cost_distress_credit", "credit_loan_repayments",
                 "income_employment", "income_benefits_state_support", "income_other"}
_TAX = {r["detailed_category"]: r for r in csv.DictReader(open(ROOT / "taxonomy" / "taxonomy.csv"))}


def _risky(leaf):
    t = _TAX.get(leaf)
    if t is None:
        return False
    return (t["general_category"] in RISK_GENERALS or t["is_priority_debt"] == "true"
            or t["is_age_restricted"] == "true" or t["is_debt_related"] == "true")


def gate():
    _, _, _, gen_of, _ = load_crosswalk()
    rows = list(csv.DictReader(open(STRINGS_CSV)))
    preds = {k: {r["merchant"]: r for r in csv.DictReader(open(PREDICTIONS[k]))} for k in MODELS}

    opus = {}
    if OPUS_PREDICTIONS.exists():
        opus = {r["merchant"]: r for r in csv.DictReader(open(OPUS_PREDICTIONS))}
        print(f"Applying {TIEBREAK_CFG['id']} tiebreak over {len(opus)} strings", file=sys.stderr)

    out, stats = [], {}
    vol = {r["merchant"]: int(r["plaid_n"]) for r in rows}
    for r in rows:
        m = r["merchant"]
        h, s = preds["haiku"].get(m, {}), preds["sonnet"].get(m, {})
        hl, sl = h.get("llm_leaf", ""), s.get("llm_leaf", "")
        sc = float(s["llm_confidence"]) if s.get("llm_confidence") else 0.0
        agree = bool(hl) and hl == sl
        if agree and hl == "unclassified_other":
            tier, leaf = "abstain_confirmed", "unclassified_other"
        elif agree and sc >= 0.9:
            tier, leaf = "auto_accept", sl
        elif agree and sc >= 0.7:
            tier, leaf = "accepted", sl
        else:
            tier, leaf = "needs_review", sl or hl
            # 2-of-3 majority with the tiebreaker model resolves the queue;
            # the tiebreaker must be IN the majority (it never rescues a
            # low-confidence haiku+sonnet pair it disagrees with)
            ol = opus.get(m, {}).get("llm_leaf", "")
            if ol and (ol == sl or ol == hl):
                if ol == "unclassified_other":
                    tier, leaf = "abstain_confirmed", "unclassified_other"
                else:
                    tier, leaf = "accepted_tiebreak", ol
            elif ol:
                # three-way leaf disagreement. Resolve by what actually matters:
                # same general category from all three -> the leaf choice is a
                # granularity quibble, accept at general level (sonnet's leaf).
                # Risk-dimension divergence (gambling/debt/income/age) -> human,
                # never a guess. Everything else -> abstain; the T6 crosswalk
                # still categorises these at runtime, we just don't override it.
                gens = {gen_of.get(x, "") for x in (hl, sl, ol)}
                if len(gens) == 1 and "" not in gens:
                    tier, leaf = "accepted_general", sl
                elif len({_risky(x) for x in (hl, sl, ol)}) > 1:
                    tier, leaf = "needs_review", sl
                else:
                    tier, leaf = "abstain_residual", "unclassified_other"
        out.append({"merchant": m, "final_leaf": leaf, "tier": tier,
                    "haiku_leaf": hl, "sonnet_leaf": sl, "sonnet_conf": sc,
                    "opus_leaf": opus.get(m, {}).get("llm_leaf", ""),
                    "general_category": gen_of.get(leaf, ""), "plaid_n": vol[m]})
        stats.setdefault(tier, [0, 0])
        stats[tier][0] += 1
        stats[tier][1] += vol[m]

    with open(LABELS_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out[0].keys()))
        w.writeheader()
        w.writerows(sorted(out, key=lambda x: -x["plaid_n"]))

    total_n = len(out)
    total_v = sum(vol.values())
    print(f"Gated {total_n} strings ({total_v} txns of volume):")
    for tier in ("auto_accept", "accepted", "accepted_tiebreak", "accepted_general",
                 "abstain_confirmed", "abstain_residual", "needs_review"):
        n, v = stats.get(tier, (0, 0))
        print(f"  {tier:18s} {n:6d} strings ({n / total_n:5.1%})   {v:8d} txns ({v / total_v:5.1%} of volume)")
    print(f"Wrote {LABELS_CSV}")


REVIEW_XLSX = OUT_DIR / "production_review.xlsx"
REVIEW_VERDICTS = ["sonnet_correct", "haiku_correct", "opus_correct", "override",
                   "unclassifiable", "context_dependent", "unsure"]


def review_sheet():
    """Workbook for the risk-divergent three-way disagreements (needs_review)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    _, _, leaves, gen_of, notes_of = load_crosswalk()
    examples_of = load_example_merchants()
    rows = [r for r in csv.DictReader(open(LABELS_CSV)) if r["tier"] == "needs_review"]
    ev = load_evidence()

    wb = Workbook()
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")

    tax = wb.create_sheet("Taxonomy")
    tax.append(["detailed_category", "general_category", "note", "example_merchants"])
    for leaf in sorted(leaves):
        tax.append([leaf, gen_of[leaf], notes_of.get(leaf, ""), ", ".join(examples_of.get(leaf, [])[:6])])
    for cell in tax[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col, width in zip("ABCD", (36, 32, 60, 60)):
        tax.column_dimensions[col].width = width
    tax.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Review"
    headers = ["merchant", "plaid_n", "pct_credit", "median_amount",
               "plaid_native_categories", "top_raw_narratives",
               "haiku_leaf", "sonnet_leaf", "opus_leaf",
               "verdict", "correct_leaf", "notes"]
    ws.append(headers)
    for r in rows:
        e = ev.get(r["merchant"], {})
        ws.append([r["merchant"], int(r["plaid_n"]),
                   e.get("pct_credit", ""), e.get("median_amount", ""),
                   e.get("cats", ""), e.get("descs", ""),
                   r["haiku_leaf"], r["sonnet_leaf"], r["opus_leaf"],
                   "", "", ""])
    n = len(rows)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=n + 1):
        for cell in row:
            cell.font = base_font
        for c in (10, 11, 12):
            row[c - 1].fill = edit_fill
    widths = {"A": 28, "B": 9, "C": 10, "D": 12, "E": 42, "F": 60,
              "G": 24, "H": 24, "I": 24, "J": 16, "K": 26, "L": 36}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:L{n + 1}"

    dv_verdict = DataValidation(type="list", formula1=f'"{",".join(REVIEW_VERDICTS)}"', allow_blank=True)
    dv_leaf = DataValidation(type="list", formula1=f"=Taxonomy!$A$2:$A${len(leaves) + 1}", allow_blank=True)
    ws.add_data_validation(dv_verdict)
    ws.add_data_validation(dv_leaf)
    dv_verdict.add(f"J2:J{n + 1}")
    dv_leaf.add(f"K2:K{n + 1}")

    ins = wb.create_sheet("Instructions", 0)
    for label_txt, text in [
        ("Production labelling -- risk-boundary review", ""),
        ("What this is", f"{n} strings where all three models disagree AND the candidate labels differ on a risk dimension (gambling / debt / income / age-restriction). These are the only strings from tranche 1 needing human eyes -- a wrong guess here is exactly what the taxonomy exists to prevent. Sorted by volume; top-down partial passes are fine."),
        ("What to edit", "Yellow columns only. verdict picks which model is right (or override + correct_leaf, unclassifiable, context_dependent). Save in place; Claude ingests it back with the same verdict semantics as before."),
    ]:
        ins.append([label_txt, text])
    for row in ins.iter_rows():
        row[0].font = Font(name="Arial", size=10, bold=True)
        row[1].font = base_font
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ins["A1"].font = Font(name="Arial", size=14, bold=True)
    ins.column_dimensions["A"].width = 34
    ins.column_dimensions["B"].width = 110

    wb.save(REVIEW_XLSX)
    print(f"Wrote {REVIEW_XLSX} ({n} rows)", file=sys.stderr)


REVIEW_COMPLETED_XLSX = OUT_DIR / "production_review_completed.xlsx"


def apply_review(path=None):
    """Fold the human verdicts from a completed review workbook back into
    production_labels.csv. Human verdicts are final: they outrank every
    model tier. Re-runnable: after a re-gate, re-apply every archived
    workbook (data/production_review_*_completed.xlsx) plus the current one."""
    from openpyxl import load_workbook
    from collections import Counter

    _, _, _, gen_of, _ = load_crosswalk()
    if path is None:
        path = REVIEW_COMPLETED_XLSX if REVIEW_COMPLETED_XLSX.exists() else REVIEW_XLSX
    path = pathlib.Path(path)
    print(f"Applying verdicts from {path}", file=sys.stderr)
    ws = load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    col = {n: hdr.index(n) for n in ("merchant", "haiku_leaf", "sonnet_leaf", "opus_leaf",
                                     "verdict", "correct_leaf", "notes")}
    resolutions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        m = row[col["merchant"]]
        if m is None:
            continue
        v = (row[col["verdict"]] or "").strip()
        if v and v not in REVIEW_VERDICTS:
            sys.exit(f"Row '{m}': verdict '{v}' not in {REVIEW_VERDICTS}")
        cl = row[col["correct_leaf"]]
        if cl and cl not in gen_of:
            sys.exit(f"Row '{m}': correct_leaf '{cl}' is not a taxonomy leaf")
        if v in ("haiku_correct", "sonnet_correct", "opus_correct"):
            resolutions[m] = (row[col[v.replace("_correct", "_leaf")]], "human_reviewed")
        elif v == "override":
            resolutions[m] = (cl, "human_reviewed")
        elif v == "unclassifiable":
            resolutions[m] = ("unclassified_other", "abstain_human")
        elif v == "context_dependent":
            resolutions[m] = ("unclassified_other", "context_dependent")
        elif v == "unsure":
            resolutions[m] = ("unclassified_other", "abstain_residual")

    rows = list(csv.DictReader(open(LABELS_CSV)))
    applied = 0
    for r in rows:
        if r["merchant"] in resolutions and r["tier"] == "needs_review":
            r["final_leaf"], r["tier"] = resolutions[r["merchant"]]
            r["general_category"] = gen_of.get(r["final_leaf"], "")
            applied += 1
    with open(LABELS_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    total_v = sum(int(r["plaid_n"]) for r in rows)
    print(f"Applied {applied} human verdicts. Final tranche distribution:")
    stats = {}
    for r in rows:
        stats.setdefault(r["tier"], [0, 0])
        stats[r["tier"]][0] += 1
        stats[r["tier"]][1] += int(r["plaid_n"])
    for tier, (n, v) in sorted(stats.items(), key=lambda kv: -kv[1][1]):
        print(f"  {tier:18s} {n:6d} strings   {v / total_v:5.1%} of volume")
    ctx = [r["merchant"] for r in rows if r["tier"] == "context_dependent"]
    if ctx:
        print(f"Context-dependent rule candidates from this tranche: {', '.join(ctx)}")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "tiebreak", "gate", "review-sheet", "apply-review"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch(int(args[1]) if len(args) > 1 else DEFAULT_N)
    elif args[0] == "label":
        model_key = args[1] if len(args) > 1 else "haiku"
        if model_key not in MODELS:
            sys.exit(f"Unknown model '{model_key}'")
        label(model_key)
    elif args[0] == "tiebreak":
        tiebreak()
    elif args[0] == "gate":
        gate()
    elif args[0] == "review-sheet":
        review_sheet()
    elif args[0] == "apply-review":
        apply_review(args[1] if len(args) > 1 else None)
