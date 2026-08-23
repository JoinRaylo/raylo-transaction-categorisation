"""Gold set v4 -- volume-weighted evaluation for the LLM/SLM categoriser
tier specifically (agreed 2026-08-23, following v3's methodology).

v3 (build_gold_v3_volume.py) sampled true-random from the ENTIRE Plaid
population to answer "what fraction of all transaction volume gets
classified correctly", scoring the deterministic pipeline end-to-end. That's
the right question for the pipeline, but most of that volume never reaches
the LLM/SLM tier at all -- it's resolved by T1-T6 rules/dictionary first.

v4 answers a different, complementary question: **given the population the
LLM/SLM tier actually sees in production (the unmatched-Plaid residual),
weighted by true transaction volume within that residual, how accurate is
it really?** This sits alongside the existing merchant-disjoint
gold_v2_slm_eval_holdout.csv (the generalization "floor") to give the
realistic "what does today's traffic actually look like" number.

Auto-resolve uses Tier A (gold_transactions_v2/v3) ONLY, never Tier B
(production_labels_tranche3) -- Tier B is itself Haiku+Sonnet+Opus-consensus
derived, and this set exists specifically to benchmark those same models,
so using Tier B to skip review would quietly reintroduce the exact
circularity this project has repeatedly had to catch and fix.

Usage:
    python src/build_gold_v4_slm_volume.py fetch
    python src/build_gold_v4_slm_volume.py label haiku
    python src/build_gold_v4_slm_volume.py label sonnet
    python src/build_gold_v4_slm_volume.py sheet
    python src/build_gold_v4_slm_volume.py apply [path]
    python src/build_gold_v4_slm_volume.py score
"""
import csv
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from build_gold_v3_volume import _established_truth  # noqa: E402 -- Tier A only, reused as-is
from gating_experiment import (  # noqa: E402
    MODELS, build_system_prompt, build_tool_schema, load_example_merchants, load_example_notes,
    build_notes_addendum, load_crosswalk,
)
from build_final_gold_v2 import TXN_ADDENDUM  # noqa: E402 -- now carries the loan-keyword fix too

SAMPLE_CSV = OUT_DIR / "gold_v4_sample.csv"
PREDICTIONS = {k: OUT_DIR / f"gold_v4_predictions_{k}.csv" for k in MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v4_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v4_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v4_slm_volume.csv"
N_PLAID = 900

UNMATCHED_TXN_QUERY = f"""
WITH eqx_vendors AS (
  SELECT DISTINCT LOWER(TRIM(VendorDescription)) AS v
  FROM `raylo-production.equifax_data.open_banking_full_dump`
  WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
),
unmatched_merchants AS (
  SELECT DISTINCT LOWER(TRIM(merchant_name)) AS m
  FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
  WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
    AND LOWER(TRIM(merchant_name)) NOT IN (SELECT v FROM eqx_vendors)
)
SELECT LOWER(TRIM(t.merchant_name)) AS merchant, t.merchant_name AS merchant_raw,
       COALESCE(t.original_description, t.transaction_name) AS description_raw,
       t.amount, IF(t.amount < 0, 'credit', 'debit') AS direction,
       t.credit_category_detailed AS native_category
FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions` t
JOIN unmatched_merchants u ON LOWER(TRIM(t.merchant_name)) = u.m
ORDER BY RAND()
LIMIT {N_PLAID}
"""


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    established = _established_truth()
    print(f"{len(established)} merchants have an established, non-conflicting Tier A gold_leaf "
          f"-- will be auto-resolved if sampled (Tier B is deliberately NOT used here)", file=sys.stderr)

    print(f"Sampling {N_PLAID} transactions, true random, from the unmatched-Plaid population "
          f"(volume-weighted within that residual)...", file=sys.stderr)
    rows = [dict(r) for r in client.query(UNMATCHED_TXN_QUERY).result()]

    all_rows = []
    for r in rows:
        m = r["merchant"]
        all_rows.append({
            "merchant": m, "merchant_raw": r["merchant_raw"],
            "description_raw": r["description_raw"] or "", "amount": r["amount"],
            "direction": r["direction"], "native_category": r["native_category"],
            "provider": "plaid",
            "established_leaf": established.get(m, ""),
            "haiku_leaf": "", "sonnet_leaf": "", "agree": "", "proposed_gold_leaf": "",
        })

    for i, r in enumerate(all_rows):
        r["row_id"] = i
    n_established = sum(1 for r in all_rows if r["established_leaf"])
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(all_rows)
    print(f"Wrote {SAMPLE_CSV}: {len(all_rows)} rows, "
          f"{n_established} ({n_established/len(all_rows):.1%}) already resolvable from established "
          f"Tier A truth, {len(all_rows) - n_established} need fresh labelling", file=sys.stderr)


def label(model_key):
    import anthropic

    cfg = MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    # Same prompt family as production_labelling.py's finalized standard (2026-08-22):
    # taxonomy + the loan-keyword bugfix + the full 375-example corpus. TXN_ADDENDUM
    # is the single-transaction-context variant (mentions native_category explicitly,
    # since these rows carry that field) but now has the SAME bugfix as TAIL_ADDENDUM.
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TXN_ADDENDUM + build_notes_addendum(load_example_notes()))
    tool = build_tool_schema(leaves)

    rows = [r for r in csv.DictReader(open(SAMPLE_CSV)) if not r["established_leaf"]]
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)

    def row_key(r):
        return r["row_id"]

    todo = [r for r in rows if row_key(r) not in predictions]
    client = anthropic.Anthropic()
    BATCH = 20
    n_batches = (len(todo) + BATCH - 1) // BATCH

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "amount", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow({"row_id": p["row_id"], "merchant": p["merchant"], "amount": p["amount"],
                            "llm_leaf": p["llm_leaf"], "llm_confidence": p["llm_confidence"]})

    def classify_batch(batch, tag, attempt=0):
        user_msg = ("Classify each of these real transactions:\n\n"
                    + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
        try:
            resp = client.messages.create(
                model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                system=system_prompt, tools=[tool],
                tool_choice={"type": "tool", "name": "submit_classifications"},
                messages=[{"role": "user", "content": user_msg}],
                timeout=90.0,
                **cfg.get("extra", {}),
            )
        except Exception as e:
            if attempt < 2:
                print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                import time; time.sleep(2 ** attempt)
                return classify_batch(batch, tag, attempt + 1)
            print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
            return {}
        tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
        if not tool_use:
            return {}
        by_idx = {j + 1: r for j, r in enumerate(batch)}
        out = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            r = by_idx.get(idx)
            if not r:
                continue
            out[row_key(r)] = {"row_id": r["row_id"], "merchant": r["merchant"], "amount": r["amount"],
                                "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
        return out

    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 10 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:04d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if row_key(r) not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:04d}_r{attempt}"))
        if num % 15 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if row_key(r) not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    haiku = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["haiku"]))} \
        if PREDICTIONS["haiku"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        if r["established_leaf"]:
            r["proposed_gold_leaf"] = r["established_leaf"]
            r["agree"] = "established"
            continue
        k = r["row_id"]
        h, s = haiku.get(k), sonnet.get(k)
        r["haiku_leaf"] = h["llm_leaf"] if h else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        if h and s and h["llm_leaf"] == s["llm_leaf"]:
            r["proposed_gold_leaf"] = h["llm_leaf"]
            r["agree"] = "yes"
        elif h and s:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "no"
        else:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "missing"

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v4 (SLM/LLM-tier volume-weighted) -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    n_established = sum(1 for r in rows if r["agree"] == "established")
    n_review = len(rows) - n_established
    ws_instr["A3"] = "What this is"
    ws_instr["B3"] = (f"A TRUE random sample of real transactions from the UNMATCHED-Plaid population "
                      f"specifically (what the LLM/SLM categoriser tier actually sees in production), "
                      f"weighted by true volume within that residual. {n_established} of {len(rows)} rows "
                      f"already have an established, non-conflicting answer from Tier A (gold_v2/v3) -- "
                      f"marked agree='established', final_leaf pre-filled, spot-check only. Deliberately NOT "
                      f"auto-resolved against Tier B (the production tranches), since that's LLM-consensus-"
                      f"derived and this set exists to benchmark those same models. {n_review} rows are "
                      f"genuinely new, needing your review, same process as v2/v3.")
    ws_instr["A4"] = "What to do"
    ws_instr["B4"] = ("Fill in `final_leaf` for every row (Taxonomy sheet has the valid list). Leave blank only "
                      "if genuinely unclassifiable.")
    ws_instr.column_dimensions["A"].width = 18
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "haiku_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf", "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["haiku_leaf"], r["sonnet_leaf"], r["agree"],
                   r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).fill = yellow
        ws.cell(row=row_idx, column=12).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 20, 20, 12, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows ({n_established} pre-resolved, "
          f"{n_disagree} genuine disagreements)", file=sys.stderr)


def apply_review(path):
    """Produces data/gold_transactions_v4_slm_volume.csv in the SAME schema as
    data/gold_v2_slm_eval_holdout.csv (merchant_raw, description_raw, amount,
    direction, gold_leaf) so it drops directly into every scoring script
    already built this session -- just point GOLD_CSV at the new file."""
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        # Plaid's raw amount is signed (negative = credit) -- the system prompt
        # promises "amount (absolute value, GBP)"; same fix as commit 910e14c.
        amount = row[idx["amount"]]
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]] or "",
            "amount": abs(float(amount)), "direction": row[idx["direction"]],
            "gold_leaf": final_leaf,
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant_raw", "description_raw", "amount", "direction", "gold_leaf"])
        w.writeheader(); w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in MODELS:
            sys.exit(f"Usage: label [{'|'.join(MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
