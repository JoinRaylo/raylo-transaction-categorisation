"""LLM merchant-labelling gating experiment (CLAUDE.md section 6).

Decides whether to invest in LLM-labelling the full 209,985-string Plaid
merchant vocabulary, or skip straight to training on Equifax's 44.7M
labelled transactions.

Method: take the 2,315 merchant strings that appear in both providers'
transaction data (measured in `analyse_provider_disagreement.py`). For each,
Equifax's own PrimaryCategoryDescription/SubCategoryDescription resolve to a
taxonomy leaf *without ever looking at the merchant string* -- that's
independent ground truth. Ask Haiku to guess the leaf from the Plaid
merchant string alone (blind to any provider category, exactly the
real-world Plaid-inference problem), then score agreement.

Ground truth excludes merchants whose modal Equifax primary is one of the
13 mechanism-override (T3) primaries -- those transactions resolve before
the merchant-classification tier ever runs in production, so testing an
LLM against them isn't a fair test of merchant-text classification.

Usage:
    python src/gating_experiment.py fetch           # BigQuery -> outputs/gating_ground_truth.csv
    python src/gating_experiment.py label [haiku|sonnet]   # LLM calls -> outputs/gating_llm_predictions*.csv
    python src/gating_experiment.py score           # scores every prediction file present + cross-model analysis
    python src/gating_experiment.py run [haiku|sonnet]     # fetch + label + score
"""
import csv
import json
import pathlib
import subprocess
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parents[1]
TAXONOMY_CSV = ROOT / "taxonomy" / "taxonomy.csv"
OUT_DIR = ROOT / "outputs"
GROUND_TRUTH_CSV = OUT_DIR / "gating_ground_truth.csv"
REPORT_MD = OUT_DIR / "gating_report.md"
CANDIDATE_GT_ERRORS_CSV = OUT_DIR / "gating_candidate_gt_errors.csv"
ADJUDICATION_XLSX = OUT_DIR / "gating_adjudication.xlsx"
ADJUDICATION_REPORT_MD = OUT_DIR / "gating_adjudication_report.md"
DICT_ADDITIONS_CSV = OUT_DIR / "gating_dictionary_additions.csv"

# Closed annotation vocabulary -- `adjudicate` parses these exact strings.
# `context_dependent` = no single merchant-level leaf is correct (the string is
# polysemous per transaction, or names different entities in the two providers);
# excluded from accuracy and reported as a T1/T2 rule candidate instead.
VERDICTS = ["llm_correct", "equifax_correct", "both_acceptable", "both_wrong", "context_dependent", "unsure"]

# Evidence pulled from BigQuery for the adjudication sheet (see review-sheet).
EQX_MIX_JSON = OUT_DIR / "gating_eqx_category_mix.json"
EQX_DESC_JSON = OUT_DIR / "gating_eqx_top_descriptions.json"
PLAID_NATIVE_JSON = OUT_DIR / "gating_plaid_native.json"

# Per-model request config. Haiku 4.5 accepts temperature=0; Sonnet 5 rejects
# non-default sampling params entirely and runs adaptive thinking by default
# (thinking tokens share max_tokens, hence the larger budget). Same prompt and
# tool schema for both, so the comparison is model-only.
MODELS = {
    "haiku": {
        "id": "claude-haiku-4-5",
        "max_tokens": 8192,
        "extra": {"temperature": 0},
        "predictions": OUT_DIR / "gating_llm_predictions.csv",
    },
    "sonnet": {
        "id": "claude-sonnet-5",
        "max_tokens": 16000,
        "extra": {},
        "predictions": OUT_DIR / "gating_llm_predictions_sonnet.csv",
    },
}
BATCH_SIZE = 100

# T3 mechanism-override primaries: leaf is determined by these regardless of
# merchant, so they'd never reach the merchant-classification tier in
# production. Excluded from the gating test set for the same reason.
MECH_PRIMARIES = {
    "Identified Salary", "Refund", "Benefits", "Welfare", "Pension Payout",
    "Tax Refund", "Cash Back", "Cash Machine", "Cash Deposit", "Interest",
    "Interests and Dividends", "Balance Transfers", "Adjustments",
}

GROUND_TRUTH_QUERY = r"""
WITH eqx AS (
  SELECT LOWER(TRIM(VendorDescription)) AS merchant,
         PrimaryCategoryDescription AS pri, SubCategoryDescription AS sub, COUNT(*) AS n
  FROM `raylo-production.equifax_data.open_banking_full_dump`
  WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != '' AND TransactionTypeId = 2
  GROUP BY 1, 2, 3
),
eqx_modal AS (
  SELECT merchant, pri, sub, n AS eqx_n
  FROM eqx
  QUALIFY ROW_NUMBER() OVER (PARTITION BY merchant ORDER BY n DESC) = 1
),
plaid AS (
  SELECT LOWER(TRIM(merchant_name)) AS merchant, COUNT(*) AS n
  FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
  WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != '' AND amount > 0
  GROUP BY 1
)
SELECT e.merchant, e.pri, e.sub, e.eqx_n, p.n AS plaid_n
FROM eqx_modal e
JOIN plaid p USING (merchant)
ORDER BY p.n DESC
"""


def load_crosswalk():
    """sub_map / pri_map: Equifax category text -> taxonomy leaf. Also returns
    per-leaf notes (from taxonomy.csv) for use as prompt disambiguation hints."""
    rows = list(csv.DictReader(open(TAXONOMY_CSV)))
    sub_map, pri_map, leaves, gen_of, notes_of = {}, {}, [], {}, {}
    for r in rows:
        leaf = r["detailed_category"]
        leaves.append(leaf)
        gen_of[leaf] = r["general_category"]
        if r.get("notes"):
            notes_of[leaf] = r["notes"].strip()
        for s in [x.strip() for x in r["equifax_source"].split(";") if x.strip()]:
            if "+" in s or "|" in s:
                continue  # compound (T2) rules aren't simple category lookups
            if s.startswith("primary:"):
                v = s[len("primary:"):].strip()
                if v != "(null)":
                    pri_map[v] = leaf
            else:
                sub_map[s] = leaf
    return sub_map, pri_map, leaves, gen_of, notes_of


def load_example_merchants():
    """leaf -> curated example merchants, from the human-reviewed merchant
    dictionary (321 entries, ~78 leaves covered). Used as disambiguation
    grounding, not as a lookup the model should rely on for other merchants."""
    by_leaf = {}
    for r in csv.DictReader(open(ROOT / "taxonomy" / "merchant_dictionary.csv")):
        by_leaf.setdefault(r["detailed_category"], []).append(r["normalised_merchant"])
    return by_leaf


def eqx_leaf(pri, sub, sub_map, pri_map):
    if sub and sub in sub_map:
        return sub_map[sub]
    if pri and pri in pri_map:
        return pri_map[pri]
    return None


def fetch_ground_truth():
    sub_map, pri_map, leaves, _, _ = load_crosswalk()
    print("Querying BigQuery for shared merchant strings (debit-only, modal)...", file=sys.stderr)
    result = subprocess.run(
        ["bq", "query", "--use_legacy_sql=false", "--format=json", "--max_rows=1000000", GROUND_TRUTH_QUERY],
        capture_output=True, text=True, check=True,
    )
    rows = json.loads(result.stdout)
    print(f"{len(rows)} shared merchants returned", file=sys.stderr)

    excluded_mech = 0
    excluded_unresolved = 0
    kept = []
    for r in rows:
        pri, sub = r.get("pri"), r.get("sub")
        if pri in MECH_PRIMARIES:
            excluded_mech += 1
            continue
        leaf = eqx_leaf(pri, sub, sub_map, pri_map)
        if leaf is None:
            excluded_unresolved += 1
            continue
        kept.append({
            "merchant": r["merchant"],
            "eqx_pri": pri or "",
            "eqx_sub": sub or "",
            "eqx_leaf": leaf,
            "eqx_n": r["eqx_n"],
            "plaid_n": r["plaid_n"],
        })

    OUT_DIR.mkdir(exist_ok=True)
    with open(GROUND_TRUTH_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant", "eqx_pri", "eqx_sub", "eqx_leaf", "eqx_n", "plaid_n"])
        w.writeheader()
        w.writerows(kept)

    print(f"Kept {len(kept)} merchants as ground truth", file=sys.stderr)
    print(f"Excluded {excluded_mech} (mechanism-override primary, not merchant-driven in production)", file=sys.stderr)
    print(f"Excluded {excluded_unresolved} (Equifax category itself unresolved to a leaf)", file=sys.stderr)
    print(f"Wrote {GROUND_TRUTH_CSV}", file=sys.stderr)


def build_system_prompt(leaves, gen_of, notes_of, examples_of):
    from collections import defaultdict
    by_gen = defaultdict(list)
    for leaf in leaves:
        by_gen[gen_of[leaf]].append(leaf)

    lines = []
    lines.append(
        "You classify UK bank-transaction merchant strings into a fixed taxonomy used for "
        "credit-risk feature engineering at a consumer credit lender. The taxonomy has 274 "
        "detailed leaf categories grouped under 29 general categories."
    )
    lines.append(
        "Some categories look similar but are deliberately kept separate because they carry "
        "different credit-risk signal -- most importantly the gambling subtypes "
        "(gambling_betting / gambling_casino / gambling_bingo / gambling_lottery / "
        "gambling_unspecified / prize_competitions), which must never be merged: lottery play "
        "and casino play are not interchangeable signals, even though both are 'gambling'. "
        "Likewise takeaway (food ordered for delivery/collection, incl. delivery apps like "
        "Deliveroo/Just Eat/Uber Eats) is distinct from restaurant_cafe (eating in). Use the "
        "example merchants below each leaf, where given, to calibrate exactly how narrow or "
        "broad a category is -- they are illustrative anchors, not an exhaustive lookup table, "
        "so classify merchants that aren't listed by the same reasoning."
    )
    lines.append("")
    lines.append("## Taxonomy (the complete, closed set of valid leaf values)")
    lines.append("Format per line: `leaf_name` -- [note, if any] (e.g. example merchants, if any)")
    for gen in sorted(by_gen):
        lines.append(f"\n### {gen}")
        for leaf in sorted(by_gen[gen]):
            note = notes_of.get(leaf, "")
            examples = examples_of.get(leaf, [])
            extra = []
            if note:
                extra.append(note)
            if examples:
                extra.append("e.g. " + ", ".join(examples[:4]))
            suffix = f" -- {'; '.join(extra)}" if extra else ""
            lines.append(f"- `{leaf}`{suffix}")
    lines.append("")
    lines.append("## Task")
    lines.append(
        "For each merchant string, choose exactly one `detailed_category` value from the list "
        "above -- never invent a category outside this list. Merchant strings are lightly "
        "cleaned raw text as they appear in UK Open Banking transaction feeds: they may include "
        "store numbers, city or branch names, 'LTD'/'LIMITED' suffixes, card-scheme prefixes, "
        "or -- for peer-to-peer transfers -- ordinary personal names or initials."
    )
    lines.append(
        "Judge each merchant using only the string itself and general knowledge of UK "
        "merchants and brands. You are not given the transaction's provider category, amount, "
        "or direction -- decide purely from what the merchant name tells you about the purpose "
        "of spending there."
    )
    lines.append(
        "Report a confidence from 0.0 to 1.0 for every classification. If a string is "
        "ambiguous, generic, or not confidently identifiable (a personal name, initials, a "
        "bare bank-transfer reference, an unfamiliar or generic-sounding business name), you "
        "MUST respond with `unclassified_other` and a low confidence rather than guess. "
        "Abstaining is strongly preferred over a low-confidence guess with any other leaf."
    )
    lines.append(
        "Return EXACTLY one result per input merchant -- never skip, merge, or add entries. "
        "In each result, echo the input's number as `index` and the merchant string exactly "
        "as given as `merchant`, so answers can be matched back to inputs."
    )
    return "\n".join(lines)


def build_tool_schema(leaves):
    return {
        "name": "submit_classifications",
        "description": "Submit taxonomy classifications for a batch of merchant strings.",
        "strict": True,
        "input_schema": {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "index": {"type": "integer", "description": "The input's 1-based number, echoed back."},
                            "merchant": {"type": "string", "description": "Echoed input merchant string, verbatim."},
                            "detailed_category": {"type": "string", "enum": sorted(leaves)},
                            "confidence": {"type": "number", "description": "0.0-1.0"},
                        },
                        "required": ["index", "merchant", "detailed_category", "confidence"],
                        "additionalProperties": False,
                    },
                }
            },
            "required": ["results"],
            "additionalProperties": False,
        },
    }


def label_all(model_key):
    import anthropic

    cfg = MODELS[model_key]
    sub_map, pri_map, leaves, gen_of, notes_of = load_crosswalk()
    examples_of = load_example_merchants()
    system_prompt = build_system_prompt(leaves, gen_of, notes_of, examples_of)
    tool = build_tool_schema(leaves)

    merchants = [r["merchant"] for r in csv.DictReader(open(GROUND_TRUTH_CSV))]
    if not merchants:
        sys.exit(f"No merchants in {GROUND_TRUTH_CSV} -- run `fetch` first")

    client = anthropic.Anthropic()
    predictions = {}
    n_batches = (len(merchants) + BATCH_SIZE - 1) // BATCH_SIZE
    raw_dir = OUT_DIR / "gating_raw" / model_key
    raw_dir.mkdir(parents=True, exist_ok=True)

    def classify_batch(batch, tag):
        """One API call over `batch`; returns {input merchant -> prediction}.
        Joins by echoed 1-based index (validated against the echoed string),
        falling back to string match -- models occasionally skip items in long
        structured arrays, so callers must check for missing keys and retry."""
        user_msg = "Classify each of these merchant strings:\n" + "\n".join(
            f"{j+1}. {m}" for j, m in enumerate(batch)
        )
        response = client.messages.create(
            model=cfg["id"],
            max_tokens=cfg["max_tokens"],
            system=[
                {"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral", "ttl": "1h"}},
            ],
            tools=[tool],
            tool_choice={"type": "tool", "name": "submit_classifications"},
            messages=[{"role": "user", "content": user_msg}],
            **cfg["extra"],
        )
        print(
            f"  [{tag}] cache_read={response.usage.cache_read_input_tokens} "
            f"cache_write={response.usage.cache_creation_input_tokens} "
            f"input={response.usage.input_tokens} output={response.usage.output_tokens}",
            file=sys.stderr,
        )
        if response.stop_reason == "max_tokens":
            print(f"  WARNING: [{tag}] truncated at max_tokens -- results incomplete", file=sys.stderr)
        tool_use = next((b for b in response.content if b.type == "tool_use"), None)
        if tool_use is None:
            print(f"  WARNING: [{tag}] no tool_use block, stop_reason={response.stop_reason}", file=sys.stderr)
            return {}
        (raw_dir / f"{tag}.json").write_text(json.dumps(tool_use.input, indent=1))

        by_string = {m.strip().lower(): m for m in batch}
        out = {}
        for r in tool_use.input.get("results", []):
            idx = r.get("index")
            echoed = (r.get("merchant") or "").strip().lower()
            merchant = None
            if isinstance(idx, int) and 1 <= idx <= len(batch):
                candidate = batch[idx - 1]
                # trust the index when the echo agrees or is missing/mangled but
                # doesn't match any other input; otherwise trust the echoed string
                if candidate.strip().lower() == echoed or echoed not in by_string:
                    merchant = candidate
            if merchant is None and echoed in by_string:
                merchant = by_string[echoed]
            if merchant is None:
                continue
            out[merchant] = {
                "detailed_category": r.get("detailed_category"),
                "confidence": r.get("confidence"),
            }
        return out

    for i in range(0, len(merchants), BATCH_SIZE):
        batch = merchants[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"[{cfg['id']}] batch {batch_num}/{n_batches} ({len(batch)} merchants)...", file=sys.stderr)
        predictions.update(classify_batch(batch, f"batch{batch_num:03d}"))

        # models sometimes silently skip entries in long structured arrays --
        # retry just the dropped merchants, up to twice
        for attempt in (1, 2):
            missing = [m for m in batch if m not in predictions]
            if not missing:
                break
            print(f"  retry {attempt}: {len(missing)} merchants dropped from batch {batch_num}", file=sys.stderr)
            predictions.update(classify_batch(missing, f"batch{batch_num:03d}_retry{attempt}"))

    out_path = cfg["predictions"]
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant", "llm_leaf", "llm_confidence"])
        w.writeheader()
        for m in merchants:
            p = predictions.get(m)
            w.writerow({
                "merchant": m,
                "llm_leaf": p["detailed_category"] if p else "",
                "llm_confidence": p["confidence"] if p else "",
            })

    missing = sum(1 for m in merchants if m not in predictions)
    print(f"Wrote {out_path} ({len(predictions)} labelled, {missing} missing)", file=sys.stderr)


def _load_scored_rows(model_key, gt, dictionary, gen_of):
    """Join one model's predictions onto ground truth; None if not yet run."""
    path = MODELS[model_key]["predictions"]
    if not path.exists():
        return None
    preds = {r["merchant"]: r for r in csv.DictReader(open(path))}
    rows = []
    for merchant, g in gt.items():
        p = preds.get(merchant, {})
        llm_leaf = p.get("llm_leaf", "")
        conf_raw = p.get("llm_confidence", "")
        conf = float(conf_raw) if conf_raw else None
        raw_match = llm_leaf == g["eqx_leaf"] if llm_leaf else False
        # Ground truth here is Equifax's category-derived leaf. Equifax's category
        # is effectively a vendor-level dictionary (modal share is ~100% for almost
        # every merchant), so disagreements are systematic convention differences,
        # not sampling noise. Where a merchant is in our human-curated dictionary
        # and the LLM agrees with THAT instead, the raw ground truth is the outlier.
        dict_leaf = dictionary.get(merchant)
        adj_match = raw_match or (bool(llm_leaf) and dict_leaf is not None and llm_leaf == dict_leaf)
        gen_match = bool(llm_leaf) and gen_of.get(llm_leaf) == gen_of.get(g["eqx_leaf"])
        rows.append({
            "merchant": merchant,
            "eqx_leaf": g["eqx_leaf"],
            "llm_leaf": llm_leaf,
            "dict_leaf": dict_leaf,
            "confidence": conf,
            "plaid_n": int(g["plaid_n"]),
            "match": raw_match,
            "adj_match": adj_match,
            "gen_match": gen_match,
            "abstained": llm_leaf == "unclassified_other",
        })
    return rows


def _model_section(model_key, rows):
    labelled = [r for r in rows if r["llm_leaf"]]
    abstained = [r for r in labelled if r["abstained"]]
    attempted = [r for r in labelled if not r["abstained"]]

    attempted_acc = sum(r["match"] for r in attempted) / len(attempted) if attempted else 0
    attempted_acc_adj = sum(r["adj_match"] for r in attempted) / len(attempted) if attempted else 0
    gen_acc = sum(r["gen_match"] for r in attempted) / len(attempted) if attempted else 0

    vol = sum(r["plaid_n"] for r in attempted)
    vol_acc = sum(r["plaid_n"] for r in attempted if r["match"]) / vol if vol else 0
    vol_gen_acc = sum(r["plaid_n"] for r in attempted if r["gen_match"]) / vol if vol else 0

    mismatch = [r for r in attempted if not r["match"]]
    near_miss = [r for r in mismatch if r["gen_match"]]
    dict_checkable = [r for r in mismatch if r["dict_leaf"] is not None]
    dict_rescued = [r for r in dict_checkable if r["llm_leaf"] == r["dict_leaf"]]

    buckets = [(0.9, 1.01), (0.7, 0.9), (0.5, 0.7), (0.0, 0.5)]
    bucket_lines = []
    for lo, hi in buckets:
        b = [r for r in attempted if r["confidence"] is not None and lo <= r["confidence"] < hi]
        if b:
            bucket_lines.append(f"  - [{lo:.1f}, {hi:.1f}): n={len(b)}, leaf accuracy={sum(r['match'] for r in b)/len(b):.1%}")

    lines = []
    lines.append(f"### {MODELS[model_key]['id']}")
    lines.append(f"- Labelled {len(labelled)}/{len(rows)}; abstained {len(abstained)} ({len(abstained)/len(labelled):.1%})")
    lines.append(f"- **Leaf accuracy (non-abstained): {attempted_acc:.1%}** raw / {attempted_acc_adj:.1%} dictionary-adjusted")
    lines.append(f"- **General-category accuracy (non-abstained): {gen_acc:.1%}** ({vol_gen_acc:.1%} volume-weighted)")
    lines.append(f"- Volume-weighted leaf accuracy: {vol_acc:.1%}")
    lines.append(f"- Mismatches: {len(mismatch)}, of which {len(near_miss)} ({len(near_miss)/len(mismatch):.0%}) are within the correct general category (leaf-granularity near-misses)")
    if dict_checkable:
        lines.append(
            f"- Of {len(dict_checkable)} mismatches checkable against our curated dictionary, "
            f"{len(dict_rescued)} ({len(dict_rescued)/len(dict_checkable):.0%}) agree with the dictionary, "
            f"i.e. the raw Equifax-derived ground truth is the outlier there"
        )
    lines.append("- Accuracy by stated confidence:")
    lines.extend(bucket_lines)
    stats = {
        "attempted": attempted, "attempted_acc": attempted_acc,
        "attempted_acc_adj": attempted_acc_adj, "gen_acc": gen_acc,
    }
    return lines, stats


def score():
    _, _, leaves, gen_of, _ = load_crosswalk()
    gt = {r["merchant"]: r for r in csv.DictReader(open(GROUND_TRUTH_CSV))}
    dictionary = {
        r["normalised_merchant"]: r["detailed_category"]
        for r in csv.DictReader(open(ROOT / "taxonomy" / "merchant_dictionary.csv"))
    }

    scored = {}
    for model_key in MODELS:
        rows = _load_scored_rows(model_key, gt, dictionary, gen_of)
        if rows is not None:
            scored[model_key] = rows
    if not scored:
        sys.exit("No prediction files found -- run `label` first")

    lines = []
    lines.append("# Gating experiment results\n")
    lines.append(f"Ground-truth merchants: {len(gt)} (strings present in both providers; debit-only; modal Equifax category)")
    lines.append("")
    lines.append("## What the ground truth actually is")
    lines.append(
        "Equifax's category is a vendor-level dictionary: for the shared merchants, the modal "
        "(primary, sub) pair covers ~100% of each merchant's transactions (median modal share 1.00; "
        "only 2% of merchants below 0.90). So this experiment measures agreement with *Equifax's "
        "merchant dictionary and its conventions* -- internally consistent, but with categorisation "
        "conventions that demonstrably differ from our taxonomy's intent in places (our own curated "
        "dictionary sides with the LLM on ~75% of checkable disputes). Leaf-level 'accuracy' below "
        "is therefore a lower bound on true label quality; the general-category figures and the "
        "cross-model agreement analysis bound it from the other side."
    )
    lines.append("")
    lines.append("## Per-model results (identical prompt, tool schema, and merchant batches)")
    all_stats = {}
    for model_key, rows in scored.items():
        section, stats = _model_section(model_key, rows)
        all_stats[model_key] = stats
        lines.extend(section)
        lines.append("")

    # Cross-model analysis: where two models independently agree with each other
    # but disagree with the Equifax-derived label, the label itself is the prime
    # suspect. Those rows go to human adjudication.
    if len(scored) >= 2:
        keys = list(scored.keys())
        a_key, b_key = keys[0], keys[1]
        a_rows = {r["merchant"]: r for r in scored[a_key]}
        b_rows = {r["merchant"]: r for r in scored[b_key]}
        both_attempted = [
            (a_rows[m], b_rows[m]) for m in a_rows
            if a_rows[m]["llm_leaf"] and b_rows[m]["llm_leaf"]
            and not a_rows[m]["abstained"] and not b_rows[m]["abstained"]
        ]
        models_agree = [(a, b) for a, b in both_attempted if a["llm_leaf"] == b["llm_leaf"]]
        agree_and_right = [(a, b) for a, b in models_agree if a["match"]]
        agree_and_wrong = [(a, b) for a, b in models_agree if not a["match"]]
        disagree = [(a, b) for a, b in both_attempted if a["llm_leaf"] != b["llm_leaf"]]

        lines.append("## Cross-model analysis")
        lines.append(f"Merchants attempted by both models: {len(both_attempted)}")
        lines.append(
            f"- Models agree with each other: {len(models_agree)} ({len(models_agree)/len(both_attempted):.1%}) -- "
            f"of these, {len(agree_and_right)} ({len(agree_and_right)/len(models_agree):.1%}) also match the Equifax-derived label"
        )
        lines.append(
            f"- **Models agree with each other but NOT with the Equifax label: {len(agree_and_wrong)}** -- "
            "two independent models converging on the same different answer makes the Equifax-derived "
            "label the prime suspect; these are exported for human adjudication"
        )
        lines.append(f"- Models disagree with each other: {len(disagree)} -- genuinely hard/ambiguous strings")
        agree_acc = len(agree_and_right) / len(models_agree) if models_agree else 0
        lines.append(
            f"- If human adjudication confirms the consensus rows, consensus-vs-truth accuracy would be "
            f"bounded between {agree_acc:.1%} (all consensus disputes are LLM errors) and "
            f"{(len(agree_and_right)+len(agree_and_wrong))/len(models_agree):.1%} (all are ground-truth errors)"
        )
        lines.append("")

        with open(CANDIDATE_GT_ERRORS_CSV, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=[
                "merchant", "eqx_leaf", "llm_consensus_leaf", "dict_leaf",
                f"{a_key}_confidence", f"{b_key}_confidence", "plaid_n",
            ])
            w.writeheader()
            for a, b in sorted(agree_and_wrong, key=lambda ab: -ab[0]["plaid_n"]):
                w.writerow({
                    "merchant": a["merchant"],
                    "eqx_leaf": a["eqx_leaf"],
                    "llm_consensus_leaf": a["llm_leaf"],
                    "dict_leaf": a["dict_leaf"] or "",
                    f"{a_key}_confidence": a["confidence"],
                    f"{b_key}_confidence": b["confidence"],
                    "plaid_n": a["plaid_n"],
                })
        lines.append(f"Candidate ground-truth errors exported to `{CANDIDATE_GT_ERRORS_CSV.name}` ({len(agree_and_wrong)} rows, sorted by Plaid volume) for human adjudication.")
        lines.append("")

    lines.append("## Verdict (CLAUDE.md §6 thresholds)")
    best_key = max(all_stats, key=lambda k: all_stats[k]["attempted_acc_adj"])
    best = all_stats[best_key]
    lines.append(
        f"Best model ({MODELS[best_key]['id']}): {best['attempted_acc']:.1%} raw / "
        f"{best['attempted_acc_adj']:.1%} dictionary-adjusted leaf accuracy, {best['gen_acc']:.1%} general-category accuracy."
    )
    if best["attempted_acc_adj"] >= 0.95:
        lines.append("**>=95% GREEN LIGHT** -- proceed to batch-label the full 209,985-string vocabulary.")
    elif best["attempted_acc_adj"] <= 0.80:
        lines.append(
            "**Below 80% at leaf level against the Equifax-derived labels -- but see the "
            "measurement-limits note above before treating this as a clean STOP.** The metric's "
            "ceiling is the quality of Equifax's own dictionary, and the candidate-ground-truth-error "
            "export quantifies how much of the shortfall is attributable to the ground truth itself. "
            "Human adjudication of that export decides the verdict."
        )
    else:
        lines.append("**AMBIGUOUS** -- between the two thresholds; needs human judgement, not an automatic call.")
    lines.append("")

    # Per-model top mismatch tables
    for model_key, rows in scored.items():
        mismatches = sorted(
            (r for r in rows if r["llm_leaf"] and not r["abstained"] and not r["match"]),
            key=lambda r: -r["plaid_n"],
        )[:30]
        lines.append(f"## Top {len(mismatches)} {MODELS[model_key]['id']} mismatches by Plaid volume")
        lines.append("| merchant | equifax (ground truth) | LLM | our dictionary | confidence | plaid_n |")
        lines.append("|---|---|---|---|---|---|")
        for r in mismatches:
            dict_note = ""
            if r["dict_leaf"] is not None:
                dict_note = f"**{r['dict_leaf']}** (LLM {'agrees' if r['adj_match'] else 'still disagrees'})"
            lines.append(f"| {r['merchant']} | {r['eqx_leaf']} | {r['llm_leaf']} | {dict_note} | {r['confidence']:.2f} | {r['plaid_n']} |")
        lines.append("")

    report = "\n".join(lines)
    REPORT_MD.write_text(report)
    print(report)
    print(f"\nWrote {REPORT_MD}", file=sys.stderr)


def build_review_sheet():
    """Turn gating_candidate_gt_errors.csv into an annotation workbook with
    dropdown-constrained verdict/correct_leaf columns, so the annotated file
    round-trips into `adjudicate` with no free-text parsing."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    _, _, leaves, gen_of, notes_of = load_crosswalk()
    examples_of = load_example_merchants()
    disputes = list(csv.DictReader(open(CANDIDATE_GT_ERRORS_CSV)))
    if not disputes:
        sys.exit(f"No rows in {CANDIDATE_GT_ERRORS_CSV} -- run `score` with both models labelled first")

    wb = Workbook()
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")

    # --- Taxonomy reference sheet (also the source range for the leaf dropdown)
    tax = wb.create_sheet("Taxonomy")
    tax.append(["detailed_category", "general_category", "note", "example_merchants"])
    for leaf in sorted(leaves):
        tax.append([leaf, gen_of[leaf], notes_of.get(leaf, ""), ", ".join(examples_of.get(leaf, [])[:6])])
    for cell in tax[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in tax.iter_rows(min_row=2):
        for cell in row:
            cell.font = base_font
    for col, width in zip("ABCD", (36, 32, 60, 60)):
        tax.column_dimensions[col].width = width
    tax.freeze_panes = "A2"

    # --- Evidence from BigQuery (optional -- columns are blank if not fetched).
    # The dispute is only judgeable in context: what mix of categories Equifax
    # actually assigned this string's transactions, what the raw bank narratives
    # look like, and what Plaid natively calls the string (Plaid being the
    # deployment target the label will be applied to).
    eqx_mix, eqx_desc, plaid_native = {}, {}, {}
    if EQX_MIX_JSON.exists():
        from collections import defaultdict
        by_m = defaultdict(list)
        for r in json.loads(EQX_MIX_JSON.read_text()):
            by_m[r["m"]].append(r)
        for m, entries in by_m.items():
            total = sum(int(r["n"]) for r in entries)
            agg = defaultdict(int)
            credit = 0
            for r in entries:
                agg[f"{r['pri']} | {r['sub'] or '(none)'}"] += int(r["n"])
                if r["t"] == "1":
                    credit += int(r["n"])
            top = sorted(agg.items(), key=lambda kv: -kv[1])[:3]
            eqx_mix[m] = (
                " · ".join(f"{cat} {cnt/total:.0%}" for cat, cnt in top),
                credit / total if total else 0,
            )
    if EQX_DESC_JSON.exists():
        for r in json.loads(EQX_DESC_JSON.read_text()):
            eqx_desc[r["m"]] = " · ".join(
                f"\"{d['value'].strip()[:60]}\" (x{d['count']})" for d in r["top_desc"][:2]
            )
    if PLAID_NATIVE_JSON.exists():
        from collections import defaultdict
        by_m = defaultdict(list)
        for r in json.loads(PLAID_NATIVE_JSON.read_text()):
            by_m[r["m"]].append(r)
        for m, entries in by_m.items():
            total = sum(int(r["n"]) for r in entries)
            top = sorted(entries, key=lambda r: -int(r["n"]))[:2]
            plaid_native[m] = " · ".join(f"{r['cat']} {int(r['n'])/total:.0%}" for r in top)

    # --- Main annotation sheet
    ws = wb.active
    ws.title = "Adjudication"
    headers = [
        "merchant", "plaid_n", "equifax_leaf", "equifax_general",
        "llm_consensus_leaf", "llm_general", "our_dictionary_leaf",
        "plaid_native_category", "eqx_category_mix", "eqx_pct_credit",
        "eqx_top_raw_descriptions", "haiku_conf", "sonnet_conf",
        "verdict", "correct_leaf", "notes",
    ]
    ws.append(headers)
    for r in disputes:
        m = r["merchant"]
        mix, pct_credit = eqx_mix.get(m, ("", ""))
        ws.append([
            m, int(r["plaid_n"]),
            r["eqx_leaf"], gen_of.get(r["eqx_leaf"], ""),
            r["llm_consensus_leaf"], gen_of.get(r["llm_consensus_leaf"], ""),
            r["dict_leaf"],
            plaid_native.get(m, ""), mix,
            round(pct_credit, 2) if pct_credit != "" else "",
            eqx_desc.get(m, ""),
            float(r["haiku_confidence"]), float(r["sonnet_confidence"]),
            "", "", "",
        ])
    n = len(disputes)
    VERDICT_COL, LEAF_COL, NOTES_COL = 14, 15, 16  # 1-based
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=n + 1):
        for cell in row:
            cell.font = base_font
        for col in (VERDICT_COL, LEAF_COL, NOTES_COL):
            row[col - 1].fill = edit_fill
    widths = {"A": 26, "B": 9, "C": 24, "D": 24, "E": 24, "F": 24, "G": 22,
              "H": 44, "I": 52, "J": 11, "K": 56, "L": 10, "M": 11,
              "N": 18, "O": 26, "P": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:P{n + 1}"

    dv_verdict = DataValidation(
        type="list", formula1=f'"{",".join(VERDICTS)}"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid verdict",
        error="Pick one of: " + ", ".join(VERDICTS),
    )
    dv_leaf = DataValidation(
        type="list", formula1=f"=Taxonomy!$A$2:$A${len(leaves) + 1}", allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid leaf",
        error="Must be a detailed_category from the Taxonomy sheet",
    )
    ws.add_data_validation(dv_verdict)
    ws.add_data_validation(dv_leaf)
    dv_verdict.add(f"N2:N{n + 1}")
    dv_leaf.add(f"O2:O{n + 1}")

    # --- Instructions sheet (with a worked example, kept OUT of the live rows)
    ins = wb.create_sheet("Instructions", 0)
    ins_rows = [
        ("Gating experiment -- human adjudication", ""),
        ("", ""),
        ("What this is", f"{n} merchants where Haiku 4.5 and Sonnet 5 independently agreed on the same taxonomy leaf, but that leaf differs from the Equifax-derived label. You decide who is right. Rows are sorted by Plaid transaction volume, so even a partial pass covers most volume -- work top-down."),
        ("The question to answer", "What is the right leaf for this string AS IT APPEARS IN PLAID DATA? Plaid is where the label will be applied. The same string can name different entities in the two providers -- e.g. Equifax's 'marks & spencer' is 100% M&S CREDIT CARD direct debits (M&S Money), while Plaid's is 99.9% department-store spend. In such cases neither side made an error; judge the Plaid string and note the entity split."),
        ("The evidence columns", "plaid_native_category = what Plaid itself calls this string (share of its transactions). eqx_category_mix = the full distribution of categories Equifax assigned this string's transactions (not just the modal one). eqx_pct_credit = share of Equifax transactions that are money IN (1.0 = all credits). eqx_top_raw_descriptions = the most common raw bank narratives behind the Equifax label -- often the fastest tell (e.g. 'M&S CREDIT CARD', 'FREEMANS - AGY')."),
        ("What to edit", "Only the three yellow columns: verdict (dropdown), correct_leaf (dropdown, only when verdict=both_wrong), notes (optional free text). Everything else is context."),
        ("", ""),
        ("verdict = llm_correct", "The models' label fits our taxonomy's intent for this merchant as it appears in Plaid; the Equifax-derived label does not."),
        ("verdict = equifax_correct", "The Equifax-derived label is right; the models are wrong."),
        ("verdict = both_acceptable", "Genuinely arguable either way (mixed-basket retailer, convention difference where we don't care). Counts in the LLM's favour, since the label wouldn't be a defect."),
        ("verdict = both_wrong", "Neither label is right -- pick the right leaf in correct_leaf."),
        ("verdict = context_dependent", "No single merchant-level leaf is correct: the string is polysemous per transaction (e.g. revolut -- Equifax's own mix is ~71% unspecified transfer / ~29% own-transfer, varying by direction) or names different entities across providers. Excluded from accuracy; reported as a candidate for a transaction-level (T1/T2) rule instead of a dictionary entry."),
        ("verdict = unsure", "Can't tell without more digging. Excluded from the corrected accuracy; reported separately."),
        ("", ""),
        ("Worked example (format only -- not a live row)", "merchant=netflix, equifax_leaf=broadband_tv_phone, llm_consensus_leaf=streaming -> verdict=llm_correct, notes='Equifax lumps Netflix with ISPs; our taxonomy separates streaming'"),
        ("", ""),
        ("When done", "Save this file in place (keep .xlsx), then run:  python src/gating_experiment.py adjudicate"),
        ("Partial passes are fine", "adjudicate reports corrected accuracy over annotated rows plus optimistic/pessimistic bounds for the rest, so you can stop when the bounds are tight enough to decide."),
        ("Bonus outputs", "llm_correct and both_wrong rows are exported as candidate merchant-dictionary (T4) additions; context_dependent rows are listed as candidates for transaction-level rules."),
    ]
    for label, text in ins_rows:
        ins.append([label, text])
    for row in ins.iter_rows():
        row[0].font = Font(name="Arial", size=10, bold=True)
        row[1].font = base_font
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ins["A1"].font = Font(name="Arial", size=14, bold=True)
    ins.column_dimensions["A"].width = 34
    ins.column_dimensions["B"].width = 110

    wb.save(ADJUDICATION_XLSX)
    print(f"Wrote {ADJUDICATION_XLSX} ({n} rows to adjudicate)", file=sys.stderr)


def adjudicate():
    """Read the annotated workbook back and compute the corrected verdict."""
    from openpyxl import load_workbook

    _, _, leaves, gen_of, _ = load_crosswalk()
    gt = {r["merchant"]: r for r in csv.DictReader(open(GROUND_TRUTH_CSV))}
    dictionary = {
        r["normalised_merchant"]: r["detailed_category"]
        for r in csv.DictReader(open(ROOT / "taxonomy" / "merchant_dictionary.csv"))
    }
    scored = {k: _load_scored_rows(k, gt, dictionary, gen_of) for k in MODELS}
    if any(v is None for v in scored.values()):
        sys.exit("Both models' predictions are required -- run `label haiku` and `label sonnet` first")

    a_rows = {r["merchant"]: r for r in scored["haiku"]}
    b_rows = {r["merchant"]: r for r in scored["sonnet"]}
    consensus = [
        a_rows[m] for m in a_rows
        if a_rows[m]["llm_leaf"] and b_rows[m]["llm_leaf"]
        and not a_rows[m]["abstained"] and not b_rows[m]["abstained"]
        and a_rows[m]["llm_leaf"] == b_rows[m]["llm_leaf"]
    ]
    consensus_right = sum(r["match"] for r in consensus)

    ws = load_workbook(ADJUDICATION_XLSX, data_only=True)["Adjudication"]
    header = [c.value for c in ws[1]]
    col = {name: header.index(name) for name in ("merchant", "verdict", "correct_leaf", "notes")}
    annotations = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant = row[col["merchant"]]
        if merchant is None:
            continue
        verdict = (row[col["verdict"]] or "").strip()
        correct_leaf = row[col["correct_leaf"]]
        notes = row[col["notes"]]
        if verdict and verdict not in VERDICTS:
            sys.exit(f"Row '{merchant}': verdict '{verdict}' is not one of {VERDICTS}")
        if correct_leaf and correct_leaf not in gen_of:
            sys.exit(f"Row '{merchant}': correct_leaf '{correct_leaf}' is not a taxonomy leaf")
        annotations[merchant] = {"verdict": verdict, "correct_leaf": correct_leaf, "notes": notes or ""}

    resolved_verdicts = {"llm_correct", "equifax_correct", "both_acceptable", "both_wrong"}
    done = {m: a for m, a in annotations.items() if a["verdict"] in resolved_verdicts}
    unsure = [m for m, a in annotations.items() if a["verdict"] == "unsure"]
    context_dep = [m for m, a in annotations.items() if a["verdict"] == "context_dependent"]
    blank = [m for m, a in annotations.items() if not a["verdict"]]
    llm_ok = [m for m, a in done.items() if a["verdict"] in ("llm_correct", "both_acceptable")]
    llm_wrong = [m for m, a in done.items() if a["verdict"] in ("equifax_correct", "both_wrong")]

    n_consensus = len(consensus)
    # context_dependent rows are resolved judgements that no single merchant-level
    # label is correct -- they leave the scoreable pool entirely rather than
    # counting for or against either side.
    n_scoreable = n_consensus - len(context_dep)
    resolved_base = consensus_right + len(llm_ok) + len(llm_wrong)
    corrected = (consensus_right + len(llm_ok)) / resolved_base if resolved_base else 0
    n_unresolved = len(unsure) + len(blank)
    pessimistic = (consensus_right + len(llm_ok)) / n_scoreable if n_scoreable else 0
    optimistic = (consensus_right + len(llm_ok) + n_unresolved) / n_scoreable if n_scoreable else 0

    from collections import Counter
    verdict_counts = Counter(a["verdict"] or "(blank)" for a in annotations.values())

    lines = []
    lines.append("# Adjudicated gating verdict\n")
    lines.append(f"Consensus rows (both models agree, non-abstained): {n_consensus}")
    lines.append(f"- Already matched the Equifax label (assumed correct): {consensus_right}")
    lines.append(f"- Disputes sent to adjudication: {len(annotations)}")
    lines.append("- Annotation breakdown: " + ", ".join(f"{k}={v}" for k, v in sorted(verdict_counts.items())))
    lines.append("")
    lines.append("## Corrected consensus accuracy (leaf level)")
    lines.append(f"- Over resolved rows: **{corrected:.1%}**")
    if context_dep:
        lines.append(f"- {len(context_dep)} context_dependent rows excluded from the scoreable pool ({n_scoreable} of {n_consensus} remain)")
    if n_unresolved:
        lines.append(f"- Bounds given {n_unresolved} unresolved (blank/unsure) rows: {pessimistic:.1%} (all wrong) to {optimistic:.1%} (all right)")
    lines.append("")
    if context_dep:
        lines.append("## Context-dependent merchants (candidates for transaction-level T1/T2 rules, not dictionary entries)")
        for m in context_dep:
            note = annotations[m]["notes"]
            lines.append(f"- {m}" + (f" -- {note}" if note else ""))
        lines.append("")
    lines.append("## Verdict (CLAUDE.md §6 thresholds, applied to the corrected figure)")
    if pessimistic >= 0.95:
        lines.append("**>=95% GREEN LIGHT** even under the pessimistic bound -- proceed with consensus-gated LLM vocabulary labelling.")
    elif optimistic < 0.80:
        lines.append("**STOP** even under the optimistic bound -- train on Equifax's 44.7M labelled transactions instead.")
    elif n_unresolved and optimistic >= 0.95 > pessimistic:
        lines.append(f"**UNDECIDED** -- the bounds straddle the 95% line; adjudicate more rows (working top-down by volume) to tighten.")
    else:
        lines.append(f"**Corrected accuracy {corrected:.1%}** -- between the thresholds; a human call on strategy, but note this applies to the consensus subset (96% of test-set volume), with non-consensus strings routed to abstain/review under any LLM plan.")
    lines.append("")

    additions = []
    for m in llm_ok:
        if annotations[m]["verdict"] == "llm_correct" and m not in dictionary:
            additions.append({"normalised_merchant": m, "detailed_category": a_rows[m]["llm_leaf"],
                              "confidence": "high", "source": "gating_adjudication", "review_status": "approved",
                              "notes": annotations[m]["notes"]})
    for m, a in done.items():
        if a["verdict"] == "both_wrong" and a["correct_leaf"] and m not in dictionary:
            additions.append({"normalised_merchant": m, "detailed_category": a["correct_leaf"],
                              "confidence": "high", "source": "gating_adjudication", "review_status": "approved",
                              "notes": a["notes"]})
    if additions:
        with open(DICT_ADDITIONS_CSV, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["normalised_merchant", "detailed_category", "confidence", "source", "review_status", "notes"])
            w.writeheader()
            w.writerows(additions)
        lines.append(f"## Merchant-dictionary additions\n{len(additions)} adjudicated merchants exported to `{DICT_ADDITIONS_CSV.name}` (T4 candidates, already human-approved by this review).")

    report = "\n".join(lines)
    ADJUDICATION_REPORT_MD.write_text(report)
    print(report)
    print(f"\nWrote {ADJUDICATION_REPORT_MD}", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    commands = {"fetch", "label", "score", "run", "review-sheet", "adjudicate"}
    if not args or args[0] not in commands:
        sys.exit(__doc__)
    cmd = args[0]
    model_key = args[1] if len(args) > 1 else "haiku"
    if model_key not in MODELS:
        sys.exit(f"Unknown model '{model_key}' -- choose from {sorted(MODELS)}")
    if cmd in ("fetch", "run"):
        fetch_ground_truth()
    if cmd in ("label", "run"):
        label_all(model_key)
    if cmd in ("score", "run"):
        score()
    if cmd == "review-sheet":
        build_review_sheet()
    if cmd == "adjudicate":
        adjudicate()
