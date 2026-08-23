"""Gold set: bespoke stratified sample of the high-risk credit/gambling leaves
(agreed with Carlos 2026-08-23, following the confusion-analysis finding).

Every existing gold set (v2/v3/v4) is volume- or breadth-weighted over the
WHOLE taxonomy, which structurally starves the categories that matter most for
credit risk: gambling subtypes and high-cost/priority credit are low-volume by
nature, so a volume-weighted sample gives them a handful of rows at best (v4
got only 21 gambling_betting rows; confusion_analysis.py found some leaves
with only 3-4 rows total across all three frontier-model benchmarks). Low n
on exactly the highest-IV, highest-fair-lending-stakes leaves is the finding
that motivated this set: it is DELIBERATELY stratified roughly EQUALLY across
34 target leaves instead of by volume, so each leaf gets enough rows to
measure real accuracy on, not just anecdotes.

Target leaves (from taxonomy.csv, all is_priority_debt=true credit leaves or
risk_flag != none, restricted to the three general categories that are
genuinely low-volume-but-high-consequence -- NOT the housing/utility priority
leaves like rent/mortgage/council_tax, which volume-weighted sampling already
covers adequately):
  - gambling (6 leaves) -- subtypes MUST be individually measurable, never
    collapsed (CLAUDE.md: combined gambling_months IV 0.0053 vs Lottery alone
    0.0498; there's a test guarding this at the taxonomy level, this gold set
    is the equivalent guard at the eval level).
  - credit_loan_repayments (17 leaves, includes bnpl)
  - high_cost_distress_credit (11 leaves)

Sourcing (two paths per leaf, since dictionary coverage varies a lot --
checked before writing this: some leaves have 100+ dictionary entries
(personal_loan_repayment 122), others have zero (overdraft_unarranged,
cash_advance, balance_transfer, account_misuse, money_management_service,
charge_card_repayment, loan_repayment_dd)):
  1. Dictionary-matched: real Plaid transactions whose merchant is already in
     taxonomy/merchant_dictionary.csv mapped to that leaf -- tests whether the
     dictionary's EXISTING belief for that leaf is actually correct.
  2. Keyword fallback: a narrative regex search, used both to top up thin
     dictionary-matched leaves and as the ONLY source for the zero-dictionary
     leaves. Keyword lists are deliberately permissive (recall over
     precision) -- human review discards false positives, but a keyword that
     never fires means genuinely zero gold coverage for that leaf.
  3. Gambling gets an ADDITIONAL broad pool from Plaid's own
     ENTERTAINMENT_CASINOS_AND_GAMBLING category regardless of dictionary
     match, specifically to surface gambling merchants the dictionary has
     never seen -- exactly where subtype-blindness would bite in production.

Two-model pair is Gemini 3.7 Flash + Sonnet 5 (Option 1, matching
production_labelling.py -- NOT Haiku, which the Option-1 refactor retired
from every live labelling path 2026-08-23).

Usage (same fetch/label/sheet/apply shape as build_gold_v4_slm_volume.py):
    python src/build_gold_risk_categories.py fetch
    python src/build_gold_risk_categories.py label gemini
    python src/build_gold_risk_categories.py label sonnet
    python src/build_gold_risk_categories.py sheet
    python src/build_gold_risk_categories.py apply [path]
"""
import csv
import json
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import (  # noqa: E402
    build_system_prompt, build_tool_schema, load_example_merchants, load_example_notes,
    build_notes_addendum, load_crosswalk,
)
from build_final_gold_v2 import TXN_ADDENDUM  # noqa: E402

SAMPLE_CSV = OUT_DIR / "gold_risk_sample.csv"
# Option 1 pair (2026-08-23): Gemini 3.7 Flash + Sonnet 5, matching
# production_labelling.py.PRODUCTION_MODELS -- not gating_experiment.MODELS,
# which is the frozen historical Haiku-vs-Sonnet gating-experiment record.
RISK_MODELS = {
    "gemini": {"backend": "gemini", "id": "gemini-3.7-flash", "extra": {}},
    "sonnet": {"backend": "anthropic", "id": "claude-sonnet-5", "max_tokens": 16000, "extra": {}},
}
PREDICTIONS = {k: OUT_DIR / f"gold_risk_predictions_{k}.csv" for k in RISK_MODELS}
REVIEW_XLSX = OUT_DIR / "gold_risk_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_risk_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_risk_categories.csv"
N_PER_LEAF = 20
N_GAMBLING_BROAD = 120  # extra undictionaried pool, subtyped by LLM+human

GAMBLING_LEAVES = ["gambling_betting", "gambling_casino", "gambling_bingo",
                   "gambling_lottery", "prize_competitions", "gambling_unspecified"]
CREDIT_LOAN_LEAVES = ["credit_card_repayment", "charge_card_repayment", "revolving_credit_repayment",
                      "personal_loan_repayment", "student_loan_repayment", "car_finance_repayment",
                      "car_lease", "hire_purchase_repayment", "retail_finance_repayment", "bnpl",
                      "credit_union_repayment", "balance_transfer", "loan_repayment_manual",
                      "loan_repayment_dd", "loan_disbursement", "financial_services_other",
                      "loan_repayment_other"]
DISTRESS_CREDIT_LEAVES = ["payday_loan", "pawnbroker", "cash_advance", "cash_advance_fee",
                          "debt_collection", "debt_enforcement", "debt_management_plan",
                          "overdraft_unarranged", "account_misuse", "credit_reporting_service",
                          "money_management_service"]
ALL_TARGET_LEAVES = GAMBLING_LEAVES + CREDIT_LOAN_LEAVES + DISTRESS_CREDIT_LEAVES

# Permissive narrative keyword fallback -- recall over precision, human review
# discards false positives. Only used to top up / substitute for dictionary
# coverage, never as ground truth itself.
KEYWORD_FALLBACK = {
    "overdraft_unarranged": r"unarranged overdraft|unauthorised overdraft|overdraft fee",
    "cash_advance": r"cash advance",
    "cash_advance_fee": r"cash advance fee",
    "balance_transfer": r"balance transfer",
    "account_misuse": r"account misuse|misuse fee",
    "money_management_service": r"debt advice|money advice|stepchange|national debtline",
    "charge_card_repayment": r"charge card|amex.*payment|american express.*payment",
    "loan_repayment_dd": r"loan.*direct debit|dd.*loan repayment",
    "debt_management_plan": r"debt management plan|\bdmp\b",
    "debt_enforcement": r"bailiff|enforcement agent|high court enforcement",
    "pawnbroker": r"pawnbroker|cash converters|the money shop|ramsdens",
    "payday_loan": r"payday loan|wonga|quickquid|sunny loans",
}

_, _, ALL_LEAVES, gen_of_all, _ = load_crosswalk()
for leaf in ALL_TARGET_LEAVES:
    assert leaf in gen_of_all, f"{leaf} not in taxonomy -- check spelling against taxonomy.csv"


def _dict_merchants_for(leaf):
    return [r["normalised_merchant"] for r in csv.DictReader(open(ROOT / "taxonomy" / "merchant_dictionary.csv"))
            if r["detailed_category"] == leaf]


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    all_rows = []
    row_id = 0
    for leaf in ALL_TARGET_LEAVES:
        merchants = _dict_merchants_for(leaf)
        clauses = []
        if merchants:
            merchant_list = ", ".join(
                '"' + m.replace("\\", "\\\\").replace('"', '\\"') + '"' for m in merchants)
            clauses.append(f"LOWER(TRIM(merchant_name)) IN ({merchant_list})")
        if leaf in KEYWORD_FALLBACK:
            pat = KEYWORD_FALLBACK[leaf].replace("\\", "\\\\")
            clauses.append(
                f"REGEXP_CONTAINS(LOWER(COALESCE(original_description, transaction_name, '')), r'{pat}')")
        if not clauses:
            print(f"  [{leaf}] no dictionary entries and no keyword fallback -- SKIPPED, zero source", file=sys.stderr)
            continue
        where = " OR ".join(clauses)
        q = f"""
        SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
               COALESCE(original_description, transaction_name) AS description_raw,
               amount, IF(amount < 0, 'credit', 'debit') AS direction,
               credit_category_detailed AS native_category
        FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
        WHERE {where}
        ORDER BY RAND()
        LIMIT {N_PER_LEAF}
        """
        rows = [dict(r) for r in client.query(q).result()]
        print(f"  [{leaf}] sourced {len(rows)} rows "
              f"({len(merchants)} dict merchants{' + keyword fallback' if leaf in KEYWORD_FALLBACK else ''})",
              file=sys.stderr)
        for r in rows:
            all_rows.append({"row_id": row_id, "target_leaf": leaf, **r, "provider": "plaid"})
            row_id += 1

    # Gambling broad pool: undictionaried native-category sample, to surface
    # merchants the dictionary has never seen -- LLM+human assign the real subtype.
    q_broad = f"""
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw,
           amount, IF(amount < 0, 'credit', 'debit') AS direction,
           credit_category_detailed AS native_category
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE credit_category_detailed = 'ENTERTAINMENT_CASINOS_AND_GAMBLING'
    ORDER BY RAND()
    LIMIT {N_GAMBLING_BROAD}
    """
    rows = [dict(r) for r in client.query(q_broad).result()]
    print(f"  [gambling_broad_pool] sourced {len(rows)} undictionaried rows", file=sys.stderr)
    for r in rows:
        all_rows.append({"row_id": row_id, "target_leaf": "gambling_broad_pool", **r, "provider": "plaid"})
        row_id += 1

    fieldnames = ["row_id", "target_leaf", "merchant", "merchant_raw", "description_raw",
                  "amount", "direction", "native_category", "provider"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_rows)
    n_leaves_covered = len({r["target_leaf"] for r in all_rows if r["target_leaf"] != "gambling_broad_pool"})
    print(f"\nWrote {SAMPLE_CSV}: {len(all_rows)} rows, "
          f"{n_leaves_covered}/{len(ALL_TARGET_LEAVES)} target leaves have >=1 source row", file=sys.stderr)


def label(model_key):
    cfg = RISK_MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TXN_ADDENDUM + build_notes_addendum(load_example_notes()))

    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)
    todo = [r for r in rows if r["row_id"] not in predictions]

    BATCH = 20

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "target_leaf", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow(p)

    if cfg["backend"] == "anthropic":
        import anthropic

        client = anthropic.Anthropic()
        tool = build_tool_schema(leaves)

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.messages.create(
                    model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                    system=system_prompt, tools=[tool],
                    tool_choice={"type": "tool", "name": "submit_classifications"},
                    messages=[{"role": "user", "content": user_msg}], timeout=90.0,
                    **cfg.get("extra", {}),
                )
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
            if not tool_use:
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in tool_use.input.get("results", []):
                r = by_idx.get(res.get("index"))
                if not r:
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"], "target_leaf": r["target_leaf"],
                                    "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
            return out

    elif cfg["backend"] == "gemini":
        import os

        from google import genai
        from google.genai import types

        client = genai.Client(api_key=os.environ["GOOGLE_API_KEY"], vertexai=False)
        leaf_list = sorted(leaves)
        index_addendum = "\n\n## Category index (output this number, not the name)\n" + "\n".join(
            f"{i + 1}. {leaf}" for i, leaf in enumerate(leaf_list))
        gemini_system = system_prompt + index_addendum
        schema = {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "index": {"type": "integer"},
                            "merchant": {"type": "string"},
                            "category_index": {"type": "integer", "minimum": 1, "maximum": len(leaf_list)},
                            "confidence": {"type": "number"},
                        },
                        "required": ["index", "merchant", "category_index", "confidence"],
                    },
                }
            },
            "required": ["results"],
        }

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.models.generate_content(
                    model=cfg["id"], contents=user_msg,
                    config=types.GenerateContentConfig(
                        system_instruction=gemini_system,
                        response_mime_type="application/json", response_schema=schema, temperature=0.0,
                    ),
                )
                data = json.loads(resp.text)
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in data.get("results", []):
                cat_idx = res.get("category_index")
                r = by_idx.get(res.get("index"))
                if not r or not (isinstance(cat_idx, int) and 1 <= cat_idx <= len(leaf_list)):
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"], "target_leaf": r["target_leaf"],
                                    "llm_leaf": leaf_list[cat_idx - 1], "llm_confidence": res.get("confidence")}
            return out

    else:
        sys.exit(f"unknown backend {cfg['backend']!r}")

    n_batches = (len(todo) + BATCH - 1) // BATCH
    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 5 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:03d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if r["row_id"] not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:03d}_r{attempt}"))
        if num % 10 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if r["row_id"] not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    gemini = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["gemini"]))} \
        if PREDICTIONS["gemini"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        k = r["row_id"]
        g, s = gemini.get(k), sonnet.get(k)
        r["gemini_leaf"] = g["llm_leaf"] if g else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        r["agree"] = "yes" if (g and s and g["llm_leaf"] == s["llm_leaf"]) else ("missing" if not (g and s) else "no")
        r["proposed_gold_leaf"] = g["llm_leaf"] if r["agree"] == "yes" else ""

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set: high-risk credit & gambling categories -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A3"] = "What this is"
    n_review = len(rows)
    ws_instr["B3"] = (
        f"A DELIBERATELY STRATIFIED sample ({n_review} rows) across the {len(ALL_TARGET_LEAVES)} gambling / "
        f"credit-loan-repayment / high-cost-distress-credit leaves, roughly {N_PER_LEAF} rows per leaf plus a "
        f"{N_GAMBLING_BROAD}-row undictionaried gambling pool -- NOT volume-weighted. Existing volume-weighted "
        f"gold sets (v3/v4) give these leaves only a handful of rows each (confusion_analysis.py found some at "
        f"n=3-4), which isn't enough to measure real accuracy on the categories that matter most for credit risk. "
        f"'target_leaf' is what sourced the row (dictionary/keyword match), not gold truth -- verify or correct it. "
        f"Rows with target_leaf='gambling_broad_pool' have NO prior leaf assumption; assign the real subtype.")
    ws_instr["A4"] = "What to do"
    ws_instr["B4"] = ("Fill in `final_leaf` for every row (Taxonomy sheet has the valid list). Pay special "
                      "attention to the gambling subtype rows -- these must never be collapsed to "
                      "gambling_unspecified when a specific subtype (betting/casino/bingo/lottery) is "
                      "identifiable; that collapse alone destroys most of the category's predictive value "
                      "(CLAUDE.md: combined gambling IV 0.0053 vs Lottery alone 0.0498).")
    ws_instr.column_dimensions["A"].width = 18
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["target_leaf", "merchant_raw", "description_raw", "amount", "direction", "native_category",
              "gemini_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf", "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["target_leaf"], r["merchant_raw"], r["description_raw"], r["amount"], r["direction"],
                   r["native_category"], r["gemini_leaf"], r["sonnet_leaf"], r["agree"],
                   r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).fill = yellow
        ws.cell(row=row_idx, column=12).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    widths = [26, 22, 45, 10, 9, 32, 20, 20, 12, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows ({n_disagree} genuine disagreements)", file=sys.stderr)


def apply_review(path):
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        amount = row[idx["amount"]]
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]] or "",
            "amount": abs(float(amount)), "direction": row[idx["direction"]],
            "gold_leaf": final_leaf, "target_leaf": row[idx["target_leaf"]],
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant_raw", "description_raw", "amount", "direction",
                                          "gold_leaf", "target_leaf"])
        w.writeheader()
        w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)
    from collections import Counter
    per_leaf = Counter(r["gold_leaf"] for r in out_rows)
    thin = [leaf for leaf in ALL_TARGET_LEAVES if per_leaf.get(leaf, 0) < 5]
    if thin:
        print(f"Still thin (<5 gold rows) after review: {thin}", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in RISK_MODELS:
            sys.exit(f"Usage: label [{'|'.join(RISK_MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
