"""Tail evaluation set for the four-field categoriser (CLAUDE.md section 6/7).

The gold set from the gating experiment covers only merchants BOTH providers
resolve -- the recognisable head. The categoriser's real population is the
~210k Plaid strings Equifax never matched: long-tail merchants, two-word
ambiguities, person-like transfer counterparties. This script builds a
stratified, human-adjudicable evaluation sample of that population.

Unlike the gating experiment there is no Equifax label at all for these
strings, so the workbook is a pure labelling exercise: both models'
context-enriched suggestions are pre-filled, the human accepts or overrides.

Usage:
    python src/build_tail_eval.py fetch            # BigQuery -> sample + evidence
    python src/build_tail_eval.py label [haiku|sonnet]
    python src/build_tail_eval.py sheet            # -> outputs/tail_eval_adjudication.xlsx
    python src/build_tail_eval.py finalise         # completed workbook -> data/gold_tail_labels.csv + report
"""
import csv
import json
import pathlib
import random
import re
import subprocess
import sys

from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from gating_experiment import (  # noqa: E402
    MODELS, ROOT, OUT_DIR, build_system_prompt, build_tool_schema,
    load_crosswalk, load_example_merchants, load_example_notes, build_notes_addendum,
)

SAMPLE_CSV = OUT_DIR / "tail_eval_sample.csv"
EVIDENCE_JSON = OUT_DIR / "tail_eval_evidence.json"
PREDICTIONS = {k: OUT_DIR / f"tail_eval_predictions_{k}.csv" for k in MODELS}
SHEET_XLSX = OUT_DIR / "tail_eval_adjudication.xlsx"
SEED = 42
BATCH = 20  # enriched inputs are much longer than bare strings

# stratum -> sample size. top_volume is deterministic (the head matters most);
# the rest are seeded random draws so the set is reproducible.
ALLOCATION = {
    "top_volume": 60,
    "volume_weighted": 60,
    "uniform_tail": 60,
    "two_word": 40,
    "short_token": 20,
    "person_like": 20,
}

VERDICTS = [
    "consensus_correct",   # models agree and their leaf is right
    "haiku_correct",       # models disagree; Haiku's leaf is right
    "sonnet_correct",      # models disagree; Sonnet's leaf is right
    "override",            # neither is right -> fill correct_leaf
    "unclassifiable",      # string genuinely carries no classifiable signal
    "context_dependent",   # no single merchant-level leaf (direction/entity varies)
    "unsure",
]

POPULATION_QUERY = r"""
WITH eqx_vendors AS (
  SELECT DISTINCT LOWER(TRIM(VendorDescription)) AS v
  FROM `raylo-production.equifax_data.open_banking_full_dump`
  WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
),
plaid AS (
  SELECT LOWER(TRIM(merchant_name)) AS m, COUNT(*) AS n
  FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
  WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
  GROUP BY 1
)
SELECT p.m, p.n
FROM plaid p LEFT JOIN eqx_vendors e ON p.m = e.v
WHERE e.v IS NULL
"""

TITLE_RE = re.compile(r"^(mr|mrs|ms|miss|dr) [a-z]")
INITIAL_RE = re.compile(r"^[a-z]\.? [a-z]+$")
TWO_WORD_RE = re.compile(r"^[a-z]+ [a-z]+$")


def classify_pattern(m):
    if TITLE_RE.match(m) or INITIAL_RE.match(m):
        return "person_like"
    if len(m) <= 4:
        return "short_token"
    if TWO_WORD_RE.match(m):
        return "two_word"
    return "long_tail"


def bq_json(query):
    result = subprocess.run(
        ["bq", "query", "--use_legacy_sql=false", "--format=json", "--max_rows=1000000", query],
        capture_output=True, text=True, check=True,
    )
    return json.loads(result.stdout)


def fetch():
    print("Querying unmatched Plaid merchant population...", file=sys.stderr)
    pop = [(r["m"], int(r["n"])) for r in bq_json(POPULATION_QUERY)]
    print(f"{len(pop)} unmatched strings, {sum(n for _, n in pop)} transactions", file=sys.stderr)

    rng = random.Random(SEED)
    by_vol = sorted(pop, key=lambda x: -x[1])
    sampled = {}  # merchant -> stratum (first stratum to claim a string keeps it)

    def take(items, k, stratum):
        for m in items:
            if len([1 for s in sampled.values() if s == stratum]) >= k:
                break
            if m not in sampled:
                sampled[m] = stratum

    take([m for m, _ in by_vol], ALLOCATION["top_volume"], "top_volume")

    remainder = [(m, n) for m, n in pop if m not in sampled]
    weights = [n for _, n in remainder]
    # volume-weighted draw without replacement (sequential, seeded)
    vw_pool = remainder[:]
    vw_weights = weights[:]
    while len([1 for s in sampled.values() if s == "volume_weighted"]) < ALLOCATION["volume_weighted"] and vw_pool:
        i = rng.choices(range(len(vw_pool)), weights=vw_weights, k=1)[0]
        m = vw_pool.pop(i)[0]
        vw_weights.pop(i)
        if m not in sampled:
            sampled[m] = "volume_weighted"

    remainder = [m for m, _ in pop if m not in sampled]
    take(rng.sample(remainder, min(len(remainder), ALLOCATION["uniform_tail"] * 3)),
         ALLOCATION["uniform_tail"], "uniform_tail")

    for stratum in ("two_word", "short_token", "person_like"):
        candidates = [m for m, _ in pop if m not in sampled and classify_pattern(m) == stratum]
        take(rng.sample(candidates, min(len(candidates), ALLOCATION[stratum] * 3)),
             ALLOCATION[stratum], stratum)

    counts = dict((m, n) for m, n in pop)
    rows = [{"merchant": m, "stratum": s, "pattern": classify_pattern(m), "plaid_n": counts[m]}
            for m, s in sampled.items()]
    rows.sort(key=lambda r: -r["plaid_n"])
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant", "stratum", "pattern", "plaid_n"])
        w.writeheader()
        w.writerows(rows)
    print(f"Sampled {len(rows)} strings -> {SAMPLE_CSV}", file=sys.stderr)

    # evidence for just the sampled strings
    def q(s):
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    in_list = ", ".join(q(r["merchant"]) for r in rows)
    evidence_query = f"""
SELECT LOWER(TRIM(merchant_name)) AS m,
       COUNT(*) AS n,
       ROUND(COUNTIF(amount < 0) / COUNT(*), 2) AS pct_credit,
       ROUND(APPROX_QUANTILES(ABS(amount), 2)[OFFSET(1)], 2) AS median_amount,
       APPROX_TOP_COUNT(credit_category_detailed, 2) AS cats,
       APPROX_TOP_COUNT(COALESCE(original_description, transaction_name), 3) AS descs
FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
  AND LOWER(TRIM(merchant_name)) IN ({in_list})
GROUP BY 1
"""
    print("Querying per-string evidence...", file=sys.stderr)
    EVIDENCE_JSON.write_text(json.dumps(bq_json(evidence_query)))
    print(f"Wrote {EVIDENCE_JSON}", file=sys.stderr)


def load_evidence():
    ev = {}
    for r in json.loads(EVIDENCE_JSON.read_text()):
        ev[r["m"]] = {
            "pct_credit": float(r["pct_credit"]),
            "median_amount": float(r["median_amount"]),
            "cats": " · ".join(f"{c['value']} {int(c['count'])}x" for c in r["cats"][:2]),
            "descs": " | ".join(f"\"{d['value'].strip()[:70]}\"" for d in r["descs"][:3]),
        }
    return ev


TAIL_ADDENDUM = (
    "\n## Additional context for this task\n"
    "Each merchant below comes with evidence aggregated from its real transactions: "
    "Plaid's native category guesses, the share of transactions that are money IN "
    "(pct_credit; 0.0 = all spending), the median amount, and the most frequent raw "
    "bank narratives. Use all of it. Direction matters: a 'merchant' whose "
    "transactions are mostly credits is usually a transfer counterparty or income "
    "source, not spending. For lenders, debt collectors and credit providers, "
    "classify by the FINANCIAL PRODUCT being paid (loan repayment, catalogue credit, "
    "debt collection), never by the merchant's trade description (a debt-litigation "
    "solicitor is debt_collection, not legal_services). Personal names and bare "
    "transfer references are transfer_p2p when the evidence supports a person-to-person "
    "payment; use unclassified_other only when the evidence is genuinely uninformative."
)


def label(model_key):
    import anthropic

    cfg = MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TAIL_ADDENDUM + build_notes_addendum(load_example_notes()))
    tool = build_tool_schema(leaves)
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    ev = load_evidence()

    client = anthropic.Anthropic()
    predictions = {}
    n_batches = (len(rows) + BATCH - 1) // BATCH

    def render(i, r):
        e = ev.get(r["merchant"], {})
        return (
            f"{i}. merchant: {r['merchant']}\n"
            f"   plaid_native_categories: {e.get('cats', 'n/a')}\n"
            f"   pct_credit: {e.get('pct_credit', 'n/a')} | median_amount_gbp: {e.get('median_amount', 'n/a')}\n"
            f"   top_raw_narratives: {e.get('descs', 'n/a')}"
        )

    def classify_batch(batch, tag):
        user_msg = "Classify each of these merchant strings using the evidence provided:\n\n" + \
            "\n".join(render(j + 1, r) for j, r in enumerate(batch))
        response = client.messages.create(
            model=cfg["id"],
            max_tokens=cfg["max_tokens"],
            system=[{"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral", "ttl": "1h"}}],
            tools=[tool],
            tool_choice={"type": "tool", "name": "submit_classifications"},
            messages=[{"role": "user", "content": user_msg}],
            **cfg["extra"],
        )
        print(f"  [{tag}] cache_read={response.usage.cache_read_input_tokens} "
              f"input={response.usage.input_tokens} output={response.usage.output_tokens}", file=sys.stderr)
        if response.stop_reason == "max_tokens":
            print(f"  WARNING: [{tag}] truncated", file=sys.stderr)
        tool_use = next((b for b in response.content if b.type == "tool_use"), None)
        if tool_use is None:
            return {}
        by_string = {r["merchant"].strip().lower(): r["merchant"] for r in batch}
        out = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            echoed = (res.get("merchant") or "").strip().lower()
            merchant = None
            if isinstance(idx, int) and 1 <= idx <= len(batch):
                candidate = batch[idx - 1]["merchant"]
                if candidate.strip().lower() == echoed or echoed not in by_string:
                    merchant = candidate
            if merchant is None and echoed in by_string:
                merchant = by_string[echoed]
            if merchant is not None:
                out[merchant] = {"leaf": res.get("detailed_category"), "conf": res.get("confidence")}
        return out

    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        num = i // BATCH + 1
        print(f"[{cfg['id']}] batch {num}/{n_batches}...", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:02d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if r["merchant"] not in predictions]
            if not missing:
                break
            print(f"  retry {attempt}: {len(missing)} dropped", file=sys.stderr)
            predictions.update(classify_batch(missing, f"b{num:02d}_r{attempt}"))

    out_path = PREDICTIONS[model_key]
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant", "llm_leaf", "llm_confidence"])
        w.writeheader()
        for r in rows:
            p = predictions.get(r["merchant"])
            w.writerow({"merchant": r["merchant"],
                        "llm_leaf": p["leaf"] if p else "",
                        "llm_confidence": p["conf"] if p else ""})
    missing = sum(1 for r in rows if r["merchant"] not in predictions)
    print(f"Wrote {out_path} ({len(predictions)} labelled, {missing} missing)", file=sys.stderr)


def sheet():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    _, _, leaves, gen_of, notes_of = load_crosswalk()
    examples_of = load_example_merchants()
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    ev = load_evidence()
    preds = {}
    for k in MODELS:
        if not PREDICTIONS[k].exists():
            sys.exit(f"Missing predictions for '{k}' -- run `label {k}` first")
        preds[k] = {r["merchant"]: r for r in csv.DictReader(open(PREDICTIONS[k]))}

    wb = Workbook()
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")

    tax = wb.create_sheet("Taxonomy")
    tax.append(["detailed_category", "general_category", "note", "example_merchants"])
    for leaf in sorted(leaves):
        tax.append([leaf, gen_of[leaf], notes_of.get(leaf, ""), ", ".join(examples_of.get(leaf, [])[:6])])
    for cell in tax[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in tax.iter_rows(min_row=2):
        for cell in row:
            cell.font = base_font
    for col, width in zip("ABCD", (36, 32, 60, 60)):
        tax.column_dimensions[col].width = width
    tax.freeze_panes = "A2"

    ws = wb.active
    ws.title = "TailEval"
    headers = ["merchant", "stratum", "plaid_n", "pct_credit", "median_amount",
               "plaid_native_categories", "top_raw_narratives",
               "haiku_leaf", "haiku_conf", "sonnet_leaf", "sonnet_conf", "models_agree",
               "verdict", "correct_leaf", "notes"]
    ws.append(headers)
    for r in rows:
        m = r["merchant"]
        e = ev.get(m, {})
        h, s = preds["haiku"].get(m, {}), preds["sonnet"].get(m, {})
        agree = bool(h.get("llm_leaf")) and h.get("llm_leaf") == s.get("llm_leaf")
        ws.append([m, r["stratum"], int(r["plaid_n"]),
                   e.get("pct_credit", ""), e.get("median_amount", ""),
                   e.get("cats", ""), e.get("descs", ""),
                   h.get("llm_leaf", ""), float(h["llm_confidence"]) if h.get("llm_confidence") else "",
                   s.get("llm_leaf", ""), float(s["llm_confidence"]) if s.get("llm_confidence") else "",
                   "YES" if agree else "no",
                   "", "", ""])
    n = len(rows)
    VERDICT_COL, LEAF_COL, NOTES_COL = 13, 14, 15
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=n + 1):
        for cell in row:
            cell.font = base_font
        for c in (VERDICT_COL, LEAF_COL, NOTES_COL):
            row[c - 1].fill = edit_fill
    widths = {"A": 28, "B": 15, "C": 8, "D": 10, "E": 12, "F": 40, "G": 60,
              "H": 24, "I": 9, "J": 24, "K": 10, "L": 11, "M": 18, "N": 26, "O": 36}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:O{n + 1}"

    dv_verdict = DataValidation(type="list", formula1=f'"{",".join(VERDICTS)}"', allow_blank=True,
                                showErrorMessage=True, errorTitle="Invalid verdict",
                                error="Pick one of: " + ", ".join(VERDICTS))
    dv_leaf = DataValidation(type="list", formula1=f"=Taxonomy!$A$2:$A${len(leaves) + 1}", allow_blank=True,
                             showErrorMessage=True, errorTitle="Invalid leaf",
                             error="Must be a detailed_category from the Taxonomy sheet")
    ws.add_data_validation(dv_verdict)
    ws.add_data_validation(dv_leaf)
    dv_verdict.add(f"M2:M{n + 1}")
    dv_leaf.add(f"N2:N{n + 1}")

    ins = wb.create_sheet("Instructions", 0)
    ins_rows = [
        ("Tail evaluation set -- human labelling", ""),
        ("", ""),
        ("What this is", f"{n} merchant strings sampled from the ~210k Plaid strings Equifax never matched -- the population the four-field categoriser must actually handle. There is NO provider ground truth here: your verdicts ARE the gold labels. Strata: top_volume (biggest unmatched strings), volume_weighted and uniform_tail (representative draws), two_word / short_token / person_like (known-hard patterns)."),
        ("The question", "What is the right taxonomy leaf for this string in Plaid data? Both models' suggestions are pre-filled with confidence, plus the evidence they saw (native categories, direction share, median amount, raw narratives). pct_credit near 1.0 = money IN (income or incoming transfer), near 0.0 = spending."),
        ("What to edit", "Only the three yellow columns: verdict, correct_leaf (for override), notes."),
        ("", ""),
        ("verdict = consensus_correct", "models_agree = YES and their shared leaf is right."),
        ("verdict = haiku_correct / sonnet_correct", "Models disagree; the named model's leaf is right."),
        ("verdict = override", "Neither model's leaf is right -- pick the right one in correct_leaf."),
        ("verdict = unclassifiable", "The string genuinely carries no classifiable signal even with the evidence (the correct behaviour is to abstain)."),
        ("verdict = context_dependent", "No single merchant-level leaf is right (direction or entity varies per transaction). Note why."),
        ("verdict = unsure", "Can't tell; excluded from the gold set."),
        ("", ""),
        ("Anchoring warning", "The suggestions come from the same models the gold set will later evaluate. Judge from the evidence first, then look at the suggestions -- overriding them is expected, not exceptional."),
        ("When done", "Save as outputs/tail_eval_adjudication_completed.xlsx (or tell Claude where it is)."),
    ]
    for label_txt, text in ins_rows:
        ins.append([label_txt, text])
    for row in ins.iter_rows():
        row[0].font = Font(name="Arial", size=10, bold=True)
        row[1].font = base_font
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ins["A1"].font = Font(name="Arial", size=14, bold=True)
    ins.column_dimensions["A"].width = 34
    ins.column_dimensions["B"].width = 110

    wb.save(SHEET_XLSX)
    agree_n = sum(1 for r in rows
                  if preds["haiku"].get(r["merchant"], {}).get("llm_leaf")
                  and preds["haiku"][r["merchant"]]["llm_leaf"] == preds["sonnet"].get(r["merchant"], {}).get("llm_leaf"))
    print(f"Wrote {SHEET_XLSX} ({n} rows; models agree on {agree_n})", file=sys.stderr)


COMPLETED_XLSX = OUT_DIR / "tail_eval_adjudication_completed.xlsx"
GOLD_TAIL_CSV = ROOT / "data" / "gold_tail_labels.csv"
TAIL_REPORT_MD = ROOT / "data" / "tail_eval_report.md"


def finalise():
    """Turn the completed workbook's verdicts into the tail gold set and score
    both models against it -- the first accuracy measurement on the actual
    deployment population (strings Equifax never matched)."""
    from openpyxl import load_workbook

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = load_workbook(COMPLETED_XLSX, data_only=True)["TailEval"]
    hdr = [c.value for c in ws[1]]
    col = {n: hdr.index(n) for n in
           ("merchant", "stratum", "plaid_n", "haiku_leaf", "sonnet_leaf", "verdict", "correct_leaf", "notes")}

    gold, excluded = [], {"context_dependent": [], "unsure": [], "blank": []}
    for row in ws.iter_rows(min_row=2, values_only=True):
        m = row[col["merchant"]]
        if m is None:
            continue
        v = (row[col["verdict"]] or "").strip()
        h, s = row[col["haiku_leaf"]], row[col["sonnet_leaf"]]
        cl = row[col["correct_leaf"]]
        if v and v not in VERDICTS:
            sys.exit(f"Row '{m}': verdict '{v}' not in {VERDICTS}")
        if cl and cl not in gen_of:
            sys.exit(f"Row '{m}': correct_leaf '{cl}' is not a taxonomy leaf")
        rec = {"merchant": m, "stratum": row[col["stratum"]], "plaid_n": row[col["plaid_n"]],
               "haiku_leaf": h or "", "sonnet_leaf": s or "", "notes": row[col["notes"]] or ""}
        if v == "consensus_correct":
            rec["gold_leaf"] = h
        elif v == "haiku_correct":
            rec["gold_leaf"] = h
        elif v == "sonnet_correct":
            rec["gold_leaf"] = s
        elif v == "override":
            rec["gold_leaf"] = cl
        elif v == "unclassifiable":
            rec["gold_leaf"] = "unclassified_other"  # abstaining IS the right answer
        else:
            excluded[v or "blank"].append(m)
            continue
        rec["gold_source"] = v
        gold.append(rec)

    GOLD_TAIL_CSV.parent.mkdir(exist_ok=True)
    with open(GOLD_TAIL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant", "gold_leaf", "gold_source", "stratum", "plaid_n", "notes"])
        w.writeheader()
        for r in sorted(gold, key=lambda r: -int(r["plaid_n"])):
            w.writerow({k: r[k] for k in ("merchant", "gold_leaf", "gold_source", "stratum", "plaid_n", "notes")})

    lines = []
    lines.append("# Tail evaluation set -- adjudicated gold labels and first model readout\n")
    lines.append(f"Sampled strings: {sum(len(v) for v in excluded.values()) + len(gold)} | "
                 f"gold-labelled: {len(gold)} | excluded: "
                 + ", ".join(f"{k}={len(v)}" for k, v in excluded.items() if v))
    lines.append("")
    lines.append("## Enriched-LLM accuracy against the tail gold set")
    lines.append("These are the models' own suggestions scored against the human verdicts on those "
                 "suggestions, so consensus/haiku/sonnet-verdict rows are correct for the named model "
                 "by construction -- the informative signal is the override/unclassifiable rate and "
                 "the per-stratum breakdown.")
    lines.append("")
    lines.append("| stratum | n | haiku leaf | sonnet leaf | haiku general | sonnet general |")
    lines.append("|---|---|---|---|---|---|")
    strata = sorted({r["stratum"] for r in gold})
    for stratum in strata + ["ALL"]:
        rows_ = gold if stratum == "ALL" else [r for r in gold if r["stratum"] == stratum]
        n = len(rows_)
        stats = {}
        for mk in ("haiku", "sonnet"):
            leaf_ok = sum(1 for r in rows_ if r[f"{mk}_leaf"] == r["gold_leaf"])
            gen_ok = sum(1 for r in rows_ if gen_of.get(r[f"{mk}_leaf"]) == gen_of.get(r["gold_leaf"]))
            stats[mk] = (leaf_ok / n if n else 0, gen_ok / n if n else 0)
        lines.append(f"| {stratum} | {n} | {stats['haiku'][0]:.0%} | {stats['sonnet'][0]:.0%} "
                     f"| {stats['haiku'][1]:.0%} | {stats['sonnet'][1]:.0%} |")
    lines.append("")
    from collections import Counter
    lines.append("## Verdict breakdown")
    all_verdicts = Counter(r["gold_source"] for r in gold)
    for k, v in excluded.items():
        if v:
            all_verdicts[k] = len(v)
    lines.append(", ".join(f"{k}={n}" for k, n in sorted(all_verdicts.items())))
    if excluded["context_dependent"]:
        lines.append("")
        lines.append("## Context-dependent strings (transaction-level rule candidates)")
        for m in excluded["context_dependent"]:
            lines.append(f"- {m}")
    report = "\n".join(lines)
    TAIL_REPORT_MD.write_text(report)
    print(report)
    print(f"\nWrote {GOLD_TAIL_CSV} and {TAIL_REPORT_MD}", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "finalise"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        model_key = args[1] if len(args) > 1 else "haiku"
        if model_key not in MODELS:
            sys.exit(f"Unknown model '{model_key}'")
        label(model_key)
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "finalise":
        finalise()
