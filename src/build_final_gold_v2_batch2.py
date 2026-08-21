"""Gold set v2, batch 2 -- a second ~1500 transactions to bring the combined
transaction-level gold set to ~3000, deliberately targeting the coverage gaps
found in batch 1 (85 of 275 taxonomy leaves had zero representation).

Composition:
  - ~150 more "already verified" rows: real transactions for merchants with an
    existing human verdict that batch 1's per-leaf cap (max 8/leaf) left out.
  - ~450 "targeted" rows: for each of the ~85 leaves missing from batch 1 (all
    of which map to a specific Equifax subcategory/primary -- Plaid's own
    categories are too coarse to reach them), pull a handful of real Equifax
    transactions from exactly that category. The models are NOT told which
    leaf was targeted -- they classify blind, same as every other row --
    targeting only decides which raw transactions get sampled.
  - Remainder (~900): broad random new sampling, same method as batch 1,
    excluding every merchant used anywhere in batch 1 or batch 2 so far.

Entirely separate files from batch 1 (which is out for review) -- nothing
here touches gold_v2_sample.csv / gold_v2_review.xlsx / their predictions.

Usage:
    python src/build_final_gold_v2_batch2.py fetch
    python src/build_final_gold_v2_batch2.py label haiku
    python src/build_final_gold_v2_batch2.py label sonnet
    python src/build_final_gold_v2_batch2.py sheet
    python src/build_final_gold_v2_batch2.py apply [path]
"""
import csv
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import MODELS, build_system_prompt, build_tool_schema, load_example_merchants, load_crosswalk  # noqa: E402
from build_final_gold_v2 import compile_already_verified, _cap_stratified_sample, TXN_ADDENDUM  # noqa: E402

BATCH1_SAMPLE = OUT_DIR / "gold_v2_sample.csv"
SAMPLE_CSV = OUT_DIR / "gold_v2_sample_batch2.csv"
PREDICTIONS = {k: OUT_DIR / f"gold_v2_predictions_batch2_{k}.csv" for k in MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v2_review_batch2.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v2_review_batch2_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v2_batch2.csv"

N_ALREADY_MORE = 150
N_PER_TARGET_LEAF = 6
N_NEW_EQUIFAX = 350
N_NEW_PLAID = 550
ROW_ID_OFFSET = 10000  # keep batch2 row_ids disjoint from batch1's 0-1499


def _norm(s):
    return (s or "").strip().lower()


def _leaf_equifax_queries(leaf, row):
    """Return a list of (pri_filter, sub_filter) tuples -- either side None means
    'any' -- derived from taxonomy.csv's equifax_source for this leaf. Handles
    plain sub-category values, 'primary:X' values, and compound T2-style rules
    ('primary:A + sub:B|C')."""
    queries = []
    for s in [x.strip() for x in row["equifax_source"].split(";") if x.strip()]:
        if "+" in s:
            parts = [p.strip() for p in s.split("+")]
            pri = sub = None
            for p in parts:
                if p.startswith("primary:"):
                    pri = p[8:].strip()
                elif p.startswith("sub:"):
                    sub = [x.strip() for x in p[4:].split("|")]
            queries.append((pri, sub))
        elif s.startswith("primary:"):
            v = s[8:].strip()
            if v != "(null)":
                queries.append((v, None))
        else:
            queries.append((None, [s]))
    return queries


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    # --- exclusion set: every merchant used anywhere in batch 1 ---
    batch1_merchants = {_norm(r["merchant"]) for r in csv.DictReader(open(BATCH1_SAMPLE))}
    print(f"{len(batch1_merchants)} merchants already used in batch 1 -- excluded", file=sys.stderr)

    already = compile_already_verified()
    remaining_already = {m: v for m, v in already.items() if m not in batch1_merchants}
    print(f"{len(remaining_already)} already-verified merchants left over from batch 1's per-leaf cap", file=sys.stderr)

    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("merchants", "STRING", sorted(remaining_already))]
    )
    plaid_sql = """
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw,
           amount, IF(amount < 0,'credit','debit') AS direction,
           credit_category_detailed AS native_category
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE LOWER(TRIM(merchant_name)) IN UNNEST(@merchants)
    QUALIFY ROW_NUMBER() OVER (PARTITION BY LOWER(TRIM(merchant_name)) ORDER BY RAND()) = 1
    """
    print("Pulling real Plaid transactions for remaining already-verified merchants...", file=sys.stderr)
    plaid_hits = {r["merchant"]: dict(r) for r in client.query(plaid_sql, job_config=job_config).result()} \
        if remaining_already else {}

    still_needed = [m for m in remaining_already if m not in plaid_hits]
    already_rows = []
    if still_needed:
        job_config2 = bigquery.QueryJobConfig(
            query_parameters=[bigquery.ArrayQueryParameter("merchants", "STRING", still_needed)]
        )
        eqx_sql = """
        SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
               Description AS description_raw, Amount AS amount,
               IF(TransactionTypeId=1,'credit','debit') AS direction,
               CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category
        FROM `raylo-production.equifax_data.open_banking_full_dump`
        WHERE LOWER(TRIM(VendorDescription)) IN UNNEST(@merchants)
        QUALIFY ROW_NUMBER() OVER (PARTITION BY LOWER(TRIM(VendorDescription)) ORDER BY RAND()) = 1
        """
        eqx_hits = {r["merchant"]: dict(r) for r in client.query(eqx_sql, job_config=job_config2).result()}
    else:
        eqx_hits = {}

    for m in remaining_already:
        hit = plaid_hits.get(m) or eqx_hits.get(m)
        if not hit:
            continue
        leaf, provenance = remaining_already[m]
        already_rows.append({
            "merchant": m, "merchant_raw": hit["merchant_raw"], "description_raw": hit["description_raw"] or "",
            "amount": hit["amount"], "direction": hit["direction"], "native_category": hit["native_category"],
            "provider": "plaid" if m in plaid_hits else "equifax",
            "proposed_gold_leaf": leaf, "source": "already_verified", "provenance": provenance,
            "haiku_leaf": "", "sonnet_leaf": "", "agree": "",
        })
    already_rows = _cap_stratified_sample(already_rows, lambda r: r["proposed_gold_leaf"], cap=6, target_n=N_ALREADY_MORE)
    print(f"Got {len(already_rows)} more already-verified rows", file=sys.stderr)

    # --- targeted rows: one real sample per missing leaf's native Equifax category ---
    tax_rows = {r["detailed_category"]: r for r in csv.DictReader(open(ROOT / "taxonomy" / "taxonomy.csv"))}
    _, _, all_leaves, gen_of, _ = load_crosswalk()
    settled_from_batch1 = {r["proposed_gold_leaf"] for r in csv.DictReader(open(BATCH1_SAMPLE)) if r.get("proposed_gold_leaf")}
    target_leaves = [l for l in all_leaves if l not in settled_from_batch1]
    print(f"{len(target_leaves)} leaves targeted for breadth (missing/unsettled after batch 1)", file=sys.stderr)

    exclude_so_far = set(batch1_merchants) | set(remaining_already)
    targeted_rows = []
    for leaf in target_leaves:
        row = tax_rows.get(leaf)
        if not row or not row["equifax_source"]:
            continue
        for pri, sub in _leaf_equifax_queries(leaf, row):
            conditions = ["TRUE"]
            params = []
            if pri:
                conditions.append("PrimaryCategoryDescription = @pri")
                params.append(bigquery.ScalarQueryParameter("pri", "STRING", pri))
            if sub:
                conditions.append("SubCategoryDescription IN UNNEST(@sub)")
                params.append(bigquery.ArrayQueryParameter("sub", "STRING", sub))
            params.append(bigquery.ArrayQueryParameter("excluded", "STRING", sorted(exclude_so_far)))
            sql = f"""
            SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
                   Description AS description_raw, Amount AS amount,
                   IF(TransactionTypeId=1,'credit','debit') AS direction,
                   CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category
            FROM `raylo-production.equifax_data.open_banking_full_dump`
            TABLESAMPLE SYSTEM (10 PERCENT)
            WHERE {' AND '.join(conditions)}
              AND VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
              AND LOWER(TRIM(VendorDescription)) NOT IN UNNEST(@excluded)
            ORDER BY RAND()
            LIMIT {N_PER_TARGET_LEAF}
            """
            job_config3 = bigquery.QueryJobConfig(query_parameters=params)
            try:
                hits = list(client.query(sql, job_config=job_config3).result())
            except Exception as e:
                print(f"  [{leaf}] query failed: {e}", file=sys.stderr)
                continue
            for r in hits:
                m = r["merchant"]
                if m in exclude_so_far:
                    continue
                exclude_so_far.add(m)
                targeted_rows.append({
                    "merchant": m, "merchant_raw": r["merchant_raw"], "description_raw": r["description_raw"] or "",
                    "amount": r["amount"], "direction": r["direction"], "native_category": r["native_category"],
                    "provider": "equifax", "proposed_gold_leaf": "", "source": "new_targeted",
                    "provenance": f"targeted_for:{leaf}", "haiku_leaf": "", "sonnet_leaf": "", "agree": "",
                })
    print(f"Got {len(targeted_rows)} targeted rows across {len(target_leaves)} target leaves", file=sys.stderr)

    # --- fill remainder with broad random sampling ---
    n_remaining = 1500 - len(already_rows) - len(targeted_rows)
    n_eqx = int(n_remaining * N_NEW_EQUIFAX / (N_NEW_EQUIFAX + N_NEW_PLAID))
    n_plaid = n_remaining - n_eqx
    job_config4 = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("excluded", "STRING", sorted(exclude_so_far))]
    )
    eqx_new_sql = f"""
    WITH base AS (
      SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
             Description AS description_raw, Amount AS amount,
             IF(TransactionTypeId=1,'credit','debit') AS direction,
             PrimaryCategoryDescription AS bucket,
             CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category,
             RAND() AS rnd
      FROM `raylo-production.equifax_data.open_banking_full_dump`
      TABLESAMPLE SYSTEM (5 PERCENT)
      WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
        AND LOWER(TRIM(VendorDescription)) NOT IN UNNEST(@excluded)
    )
    SELECT * EXCEPT(rnd) FROM base
    QUALIFY ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY rnd) <= 15
    ORDER BY RAND()
    LIMIT {n_eqx}
    """
    print(f"Sampling {n_eqx} more new Equifax transactions...", file=sys.stderr)
    eqx_new = [dict(r) for r in client.query(eqx_new_sql, job_config=job_config4).result()]

    plaid_new_sql = f"""
    WITH base AS (
      SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
             COALESCE(original_description, transaction_name) AS description_raw,
             amount, IF(amount < 0,'credit','debit') AS direction,
             credit_category_detailed AS bucket, credit_category_detailed AS native_category,
             RAND() AS rnd
      FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
      WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
        AND LOWER(TRIM(merchant_name)) NOT IN UNNEST(@excluded)
    )
    SELECT * EXCEPT(rnd) FROM base
    QUALIFY ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY rnd) <= 15
    ORDER BY RAND()
    LIMIT {n_plaid}
    """
    print(f"Sampling {n_plaid} more new Plaid transactions...", file=sys.stderr)
    plaid_new = [dict(r) for r in client.query(plaid_new_sql, job_config=job_config4).result()]

    broad_rows = []
    for r in eqx_new:
        broad_rows.append({"merchant": r["merchant"], "merchant_raw": r["merchant_raw"],
                            "description_raw": r["description_raw"] or "", "amount": r["amount"],
                            "direction": r["direction"], "native_category": r["native_category"],
                            "provider": "equifax", "proposed_gold_leaf": "", "source": "new",
                            "provenance": "", "haiku_leaf": "", "sonnet_leaf": "", "agree": ""})
    for r in plaid_new:
        broad_rows.append({"merchant": r["merchant"], "merchant_raw": r["merchant_raw"],
                            "description_raw": r["description_raw"] or "", "amount": r["amount"],
                            "direction": r["direction"], "native_category": r["native_category"],
                            "provider": "plaid", "proposed_gold_leaf": "", "source": "new",
                            "provenance": "", "haiku_leaf": "", "sonnet_leaf": "", "agree": ""})

    all_rows = already_rows + targeted_rows + broad_rows
    for i, r in enumerate(all_rows):
        r["row_id"] = ROW_ID_OFFSET + i
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(all_rows)
    print(f"Wrote {SAMPLE_CSV}: {len(already_rows)} already-verified + {len(targeted_rows)} targeted "
          f"+ {len(broad_rows)} broad new = {len(all_rows)} total", file=sys.stderr)


def label(model_key):
    import anthropic

    cfg = MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = build_system_prompt(leaves, gen_of, notes_of, load_example_merchants()) + TXN_ADDENDUM
    tool = build_tool_schema(leaves)

    rows = [r for r in csv.DictReader(open(SAMPLE_CSV)) if r["source"] in ("new", "new_targeted")]
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)

    def row_key(r):
        return r["row_id"]

    todo = [r for r in rows if row_key(r) not in predictions]
    client = anthropic.Anthropic()
    BATCH = 20
    n_batches = (len(todo) + BATCH - 1) // BATCH

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "amount", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow({"row_id": p["row_id"], "merchant": p["merchant"], "amount": p["amount"],
                            "llm_leaf": p["llm_leaf"], "llm_confidence": p["llm_confidence"]})

    def classify_batch(batch, tag, attempt=0):
        user_msg = ("Classify each of these real transactions:\n\n"
                    + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
        try:
            resp = client.messages.create(
                model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                system=system_prompt, tools=[tool],
                tool_choice={"type": "tool", "name": "submit_classifications"},
                messages=[{"role": "user", "content": user_msg}],
                timeout=90.0,
                **cfg.get("extra", {}),
            )
        except Exception as e:
            if attempt < 2:
                print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                import time; time.sleep(2 ** attempt)
                return classify_batch(batch, tag, attempt + 1)
            print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
            return {}
        tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
        if not tool_use:
            return {}
        by_idx = {j + 1: r for j, r in enumerate(batch)}
        out = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            r = by_idx.get(idx)
            if not r:
                continue
            out[row_key(r)] = {"row_id": r["row_id"], "merchant": r["merchant"], "amount": r["amount"],
                                "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
        return out

    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 10 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:04d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if row_key(r) not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:04d}_r{attempt}"))
        if num % 15 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if row_key(r) not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    haiku = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["haiku"]))} \
        if PREDICTIONS["haiku"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        if r["source"] not in ("new", "new_targeted"):
            continue
        k = r["row_id"]
        h, s = haiku.get(k), sonnet.get(k)
        r["haiku_leaf"] = h["llm_leaf"] if h else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        if h and s and h["llm_leaf"] == s["llm_leaf"]:
            r["proposed_gold_leaf"] = h["llm_leaf"]
            r["agree"] = "yes"
        elif h and s:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "no"
        else:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "missing"

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v2, batch 2 -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A3"] = "What this is"
    n_already = sum(1 for r in rows if r["source"] == "already_verified")
    n_targeted = sum(1 for r in rows if r["source"] == "new_targeted")
    n_new = sum(1 for r in rows if r["source"] == "new")
    ws_instr["B3"] = (f"Second batch of {len(rows)} transactions, bringing the combined gold set to ~3000. "
                      f"{n_already} rows already have a real human verdict (spot-check only). {n_targeted} rows "
                      f"were deliberately sourced to fill leaf categories that batch 1 had zero examples of "
                      f"(see the provenance column, 'targeted_for:<leaf>') -- the models were NOT told which "
                      f"leaf was targeted, they classified blind same as every other row, so their proposal is "
                      f"a genuine independent guess, not a confirmation of the target. {n_new} rows are broad "
                      f"new random sampling, same method as batch 1.")
    ws_instr["A4"] = "What to do"
    ws_instr["B4"] = ("Same as batch 1: fill in `final_leaf` for every row with the category YOU believe is "
                      "correct (Taxonomy sheet has the full valid list). Leave blank only if genuinely "
                      "unclassifiable. Add `notes` for anything surprising. For 'targeted_for' rows, it's "
                      "fine (and expected) if the correct answer ISN'T the leaf that was targeted -- the "
                      "targeting only chose which transaction to sample, not the answer.")
    ws_instr.column_dimensions["A"].width = 18
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "source", "provenance", "haiku_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf",
              "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["source"], r["provenance"], r["haiku_leaf"], r["sonnet_leaf"],
                   r["agree"], r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=13).fill = yellow
        ws.cell(row=row_idx, column=14).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"M2:M{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 12, 30, 20, 20, 7, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKLMN", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows, {n_disagree} model disagreements need real judgment", file=sys.stderr)


def apply_review(path):
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]],
            "amount": row[idx["amount"]], "direction": row[idx["direction"]],
            "provider": row[idx["provider"]], "native_category": row[idx["native_category"]],
            "source": row[idx["source"]], "provenance": row[idx["provenance"]],
            "haiku_leaf": row[idx["haiku_leaf"]], "sonnet_leaf": row[idx["sonnet_leaf"]],
            "gold_leaf": final_leaf, "notes": row[idx["notes"]] or "",
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader(); w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in MODELS:
            sys.exit(f"Usage: label [{'|'.join(MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
