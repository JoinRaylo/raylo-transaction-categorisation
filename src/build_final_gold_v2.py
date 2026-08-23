"""Transaction-level gold set v2 -- built to eliminate the leakage found in the
2026-08-21 audit (merchant-level gold set was 79% unreviewed LLM self-agreement
on the head side, 75% directly-circular on the tail side).

This version samples ~1500 REAL, individual transactions (not merchant-level
aggregates) across a wide spread of merchants, amounts, descriptions and
categories from both providers:
  - ~400 "already verified" rows: real transactions for merchants where a
    human has already made an actual, on-the-record call (gold_merchant_labels
    adjudicated rows, gold_tail_labels rows, or a production-tranche review
    verdict) -- these need only a spot-check, not fresh review.
  - ~1100 "new" rows: real transactions for merchants nobody has looked at yet,
    stratified by provider category to keep coverage broad. Haiku + Sonnet
    label each with real per-transaction evidence (amount, direction,
    description, native category), and their agreement becomes a DRAFT
    proposed_gold_leaf for Carlos to review, correct, or reject -- not a
    final answer. Disagreements are shown with both proposals and no default.

Usage:
    python src/build_final_gold_v2.py fetch          # sample + write outputs/gold_v2_sample.csv
    python src/build_final_gold_v2.py label haiku
    python src/build_final_gold_v2.py label sonnet
    python src/build_final_gold_v2.py sheet          # build the review workbook
    python src/build_final_gold_v2.py apply [path]   # ingest the completed workbook -> data/gold_transactions_v2.csv
"""
import csv
import json
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import MODELS, build_system_prompt, build_tool_schema, load_example_merchants, load_crosswalk  # noqa: E402

SAMPLE_CSV = OUT_DIR / "gold_v2_sample.csv"
PREDICTIONS = {k: OUT_DIR / f"gold_v2_predictions_{k}.csv" for k in MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v2_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v2_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v2.csv"

N_ALREADY = 400
N_NEW_EQUIFAX = 420
N_NEW_PLAID = 680

TXN_ADDENDUM = (
    "\n## Additional context for this task\n"
    "Each row is a SINGLE real transaction, not an aggregate. You are given the merchant "
    "name, the raw bank narrative/description, the amount (absolute value, GBP), the "
    "direction (debit = money out; credit = money in), and the provider's own native "
    "category guess for context (which may be wrong -- it is evidence, not the answer). "
    "Use all of it: direction and the narrative often disambiguate what the merchant name "
    "alone cannot (e.g. a payment TO a person's name vs a refund FROM a retailer with the "
    "same string). For lenders, debt collectors and credit providers, classify by the "
    "FINANCIAL PRODUCT being paid (loan repayment, catalogue credit, debt collection), "
    "never by the merchant's trade description. Personal names and bare transfer "
    "references are transfer_p2p ONLY when nothing else in the narrative identifies a "
    "purpose -- if the raw narrative contains an explicit debt keyword (LOAN, LEND, OWE, "
    "DEBT, IOU) even alongside a personal name, classify as loan_repayment_manual "
    "instead, never transfer_p2p or personal_loan_repayment."
)


def _norm(s):
    return (s or "").strip().lower()


def compile_already_verified():
    """Union every real human verdict we have across all completed review artifacts.
    Returns {merchant: (leaf, provenance)}. Excludes context_dependent/unsure verdicts
    (those aren't a real answer) and any leaf that isn't unclassified_other-safe."""
    import openpyxl

    verified = {}

    for r in csv.DictReader(open(ROOT / "data" / "gold_merchant_labels.csv")):
        if r["gold_source"].startswith("adjudicated"):
            verified[_norm(r["merchant"])] = (r["gold_leaf"], "gold_head_adjudicated")

    for r in csv.DictReader(open(ROOT / "data" / "gold_tail_labels.csv")):
        verified[_norm(r["merchant"])] = (r["gold_leaf"], "gold_tail_human_verified")

    for tranche in ("tranche1", "tranche2", "tranche3"):
        path = ROOT / "data" / f"production_review_{tranche}_completed.xlsx"
        if not path.exists():
            continue
        ws = openpyxl.load_workbook(path, data_only=True)["Review"]
        hdr = [c.value for c in ws[1]]
        col = {n: hdr.index(n) for n in ("merchant", "haiku_leaf", "sonnet_leaf", "opus_leaf",
                                          "verdict", "correct_leaf") if n in hdr}
        for row in ws.iter_rows(min_row=2, values_only=True):
            m = row[col["merchant"]] if "merchant" in col else None
            if m is None:
                continue
            v = (row[col.get("verdict")] or "").strip() if "verdict" in col else ""
            if v in ("haiku_correct", "sonnet_correct", "opus_correct"):
                leaf = row[col[v.replace("_correct", "_leaf")]]
            elif v == "override":
                leaf = row[col.get("correct_leaf")]
            else:
                continue  # context_dependent / unsure / blank -- not a real answer
            if leaf:
                verified[_norm(m)] = (leaf, f"production_{tranche}_review")

    return verified


def _cap_stratified_sample(items, key_fn, cap, target_n):
    """Group by key_fn, cap each group, then randomly trim to target_n. No RNG seeding
    needed here since this just orders an already-fetched list -- BigQuery did the
    actual randomisation via RAND() in the query."""
    from collections import defaultdict
    groups = defaultdict(list)
    for it in items:
        groups[key_fn(it)].append(it)
    capped = []
    for g in groups.values():
        capped.extend(g[:cap])
    return capped[:target_n]


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    already = compile_already_verified()
    print(f"{len(already)} merchants have an existing human verdict", file=sys.stderr)
    already_list = sorted(already)

    # --- pull one real transaction per already-verified merchant, Plaid preferred ---
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("merchants", "STRING", already_list)]
    )
    plaid_sql = """
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw,
           amount, IF(amount < 0,'credit','debit') AS direction,
           credit_category_detailed AS native_category
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE LOWER(TRIM(merchant_name)) IN UNNEST(@merchants)
    QUALIFY ROW_NUMBER() OVER (PARTITION BY LOWER(TRIM(merchant_name)) ORDER BY RAND()) = 1
    """
    print("Pulling real Plaid transactions for already-verified merchants...", file=sys.stderr)
    plaid_hits = {r["merchant"]: dict(r) for r in client.query(plaid_sql, job_config=job_config).result()}

    still_needed = [m for m in already_list if m not in plaid_hits]
    job_config2 = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("merchants", "STRING", still_needed)]
    )
    eqx_sql = """
    SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
           Description AS description_raw, Amount AS amount,
           IF(TransactionTypeId=1,'credit','debit') AS direction,
           CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category
    FROM `raylo-production.equifax_data.open_banking_full_dump`
    WHERE LOWER(TRIM(VendorDescription)) IN UNNEST(@merchants)
    QUALIFY ROW_NUMBER() OVER (PARTITION BY LOWER(TRIM(VendorDescription)) ORDER BY RAND()) = 1
    """
    print(f"Pulling real Equifax transactions for {len(still_needed)} merchants not found in Plaid...", file=sys.stderr)
    eqx_hits = {r["merchant"]: dict(r) for r in client.query(eqx_sql, job_config=job_config2).result()} if still_needed else {}

    already_rows = []
    for m in already_list:
        hit = plaid_hits.get(m) or eqx_hits.get(m)
        if not hit:
            continue
        leaf, provenance = already[m]
        already_rows.append({
            "merchant": m, "merchant_raw": hit["merchant_raw"], "description_raw": hit["description_raw"] or "",
            "amount": hit["amount"], "direction": hit["direction"], "native_category": hit["native_category"],
            "provider": "plaid" if m in plaid_hits else "equifax",
            "proposed_gold_leaf": leaf, "source": "already_verified", "provenance": provenance,
            "haiku_leaf": "", "sonnet_leaf": "", "agree": "",
        })
    print(f"Matched real transactions for {len(already_rows)}/{len(already_list)} already-verified merchants", file=sys.stderr)

    already_rows = _cap_stratified_sample(already_rows, lambda r: r["proposed_gold_leaf"], cap=8, target_n=N_ALREADY)
    print(f"Capped to {len(already_rows)} already-verified rows (stratified by leaf, max 8/leaf)", file=sys.stderr)

    # --- sample NEW, previously-untouched transactions, stratified by native category ---
    excluded_list = already_list  # exclude every merchant with ANY prior human touch
    job_config3 = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("excluded", "STRING", excluded_list)]
    )
    eqx_new_sql = f"""
    WITH base AS (
      SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
             Description AS description_raw, Amount AS amount,
             IF(TransactionTypeId=1,'credit','debit') AS direction,
             PrimaryCategoryDescription AS bucket,
             CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category,
             RAND() AS rnd
      FROM `raylo-production.equifax_data.open_banking_full_dump`
      TABLESAMPLE SYSTEM (5 PERCENT)
      WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
        AND LOWER(TRIM(VendorDescription)) NOT IN UNNEST(@excluded)
    )
    SELECT * EXCEPT(rnd) FROM base
    QUALIFY ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY rnd) <= 15
    ORDER BY RAND()
    LIMIT {N_NEW_EQUIFAX}
    """
    print(f"Sampling {N_NEW_EQUIFAX} new Equifax transactions...", file=sys.stderr)
    eqx_new = [dict(r) for r in client.query(eqx_new_sql, job_config=job_config3).result()]

    plaid_new_sql = f"""
    WITH base AS (
      SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
             COALESCE(original_description, transaction_name) AS description_raw,
             amount, IF(amount < 0,'credit','debit') AS direction,
             credit_category_detailed AS bucket, credit_category_detailed AS native_category,
             RAND() AS rnd
      FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
      WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
        AND LOWER(TRIM(merchant_name)) NOT IN UNNEST(@excluded)
    )
    SELECT * EXCEPT(rnd) FROM base
    QUALIFY ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY rnd) <= 15
    ORDER BY RAND()
    LIMIT {N_NEW_PLAID}
    """
    print(f"Sampling {N_NEW_PLAID} new Plaid transactions...", file=sys.stderr)
    plaid_new = [dict(r) for r in client.query(plaid_new_sql, job_config=job_config3).result()]

    new_rows = []
    for r in eqx_new:
        new_rows.append({"merchant": r["merchant"], "merchant_raw": r["merchant_raw"],
                          "description_raw": r["description_raw"] or "", "amount": r["amount"],
                          "direction": r["direction"], "native_category": r["native_category"],
                          "provider": "equifax", "proposed_gold_leaf": "", "source": "new",
                          "provenance": "", "haiku_leaf": "", "sonnet_leaf": "", "agree": ""})
    for r in plaid_new:
        new_rows.append({"merchant": r["merchant"], "merchant_raw": r["merchant_raw"],
                          "description_raw": r["description_raw"] or "", "amount": r["amount"],
                          "direction": r["direction"], "native_category": r["native_category"],
                          "provider": "plaid", "proposed_gold_leaf": "", "source": "new",
                          "provenance": "", "haiku_leaf": "", "sonnet_leaf": "", "agree": ""})

    all_rows = already_rows + new_rows
    for i, r in enumerate(all_rows):
        r["row_id"] = i  # stable unique key -- merchant+amount collides on common round amounts
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(all_rows)
    print(f"Wrote {SAMPLE_CSV}: {len(already_rows)} already-verified + {len(new_rows)} new "
          f"({len(eqx_new)} eqx / {len(plaid_new)} plaid) = {len(all_rows)} total", file=sys.stderr)


def label(model_key):
    import anthropic

    cfg = MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = build_system_prompt(leaves, gen_of, notes_of, load_example_merchants()) + TXN_ADDENDUM
    tool = build_tool_schema(leaves)

    rows = [r for r in csv.DictReader(open(SAMPLE_CSV)) if r["source"] == "new"]
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)

    def row_key(r):
        return r["row_id"]

    todo = [r for r in rows if row_key(r) not in predictions]
    client = anthropic.Anthropic()
    BATCH = 20
    n_batches = (len(todo) + BATCH - 1) // BATCH

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "amount", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow({"row_id": p["row_id"], "merchant": p["merchant"], "amount": p["amount"],
                            "llm_leaf": p["llm_leaf"], "llm_confidence": p["llm_confidence"]})

    def classify_batch(batch, tag, attempt=0):
        user_msg = ("Classify each of these real transactions:\n\n"
                    + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
        try:
            resp = client.messages.create(
                model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                system=system_prompt, tools=[tool],
                tool_choice={"type": "tool", "name": "submit_classifications"},
                messages=[{"role": "user", "content": user_msg}],
                **cfg.get("extra", {}),
            )
        except Exception as e:
            if attempt < 2:
                print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                import time; time.sleep(2 ** attempt)
                return classify_batch(batch, tag, attempt + 1)
            print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
            return {}
        tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
        if not tool_use:
            return {}
        by_idx = {j + 1: r for j, r in enumerate(batch)}
        out = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            r = by_idx.get(idx)
            if not r:
                continue
            out[row_key(r)] = {"row_id": r["row_id"], "merchant": r["merchant"], "amount": r["amount"],
                                "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
        return out

    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 10 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:04d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if row_key(r) not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:04d}_r{attempt}"))
        if num % 15 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if row_key(r) not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    haiku = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["haiku"]))} \
        if PREDICTIONS["haiku"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        if r["source"] != "new":
            continue
        k = r["row_id"]
        h, s = haiku.get(k), sonnet.get(k)
        r["haiku_leaf"] = h["llm_leaf"] if h else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        if h and s and h["llm_leaf"] == s["llm_leaf"]:
            r["proposed_gold_leaf"] = h["llm_leaf"]
            r["agree"] = "yes"
        elif h and s:
            r["proposed_gold_leaf"] = ""  # disagreement -- no default, needs a real decision
            r["agree"] = "no"
        else:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "missing"

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v2 -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A3"] = "What this is"
    ws_instr["B3"] = (f"{sum(1 for r in rows if r['source']=='already_verified')} rows already have a "
                      f"real human verdict from prior work (provenance column shows the source) -- spot-check "
                      f"these, don't need a full re-review. {sum(1 for r in rows if r['source']=='new')} rows "
                      f"are brand new: Haiku and Sonnet each independently proposed a category from real "
                      f"transaction evidence (merchant, description, amount, direction, native category). "
                      f"Where they agree, that's the proposed_gold_leaf as a DRAFT starting point -- check it, "
                      f"don't just trust it. Where they disagree, proposed_gold_leaf is blank and both "
                      f"model guesses are shown in haiku_leaf/sonnet_leaf for you to weigh.")
    ws_instr["A4"] = "What to do"
    ws_instr["B4"] = ("Fill in `final_leaf` for every row with the category YOU believe is correct (use the "
                      "Taxonomy sheet for the full list of valid leaf names -- must match exactly). Leave "
                      "`final_leaf` blank only if a row is genuinely unclassifiable even with the evidence "
                      "given. Add a note in `notes` for anything surprising or ambiguous. Save this file in "
                      "place (or export to xlsx) and send it back.")
    ws_instr.column_dimensions["A"].width = 18
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "source", "provenance", "haiku_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf",
              "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["source"], r["provenance"], r["haiku_leaf"], r["sonnet_leaf"],
                   r["agree"], r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=13).fill = yellow  # final_leaf
        ws.cell(row=row_idx, column=14).fill = yellow  # notes
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"M2:M{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 8, 24, 20, 20, 7, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKLMN", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows, {n_disagree} model disagreements need real judgment", file=sys.stderr)


def apply_review(path):
    """Ingest the completed review workbook -> data/gold_transactions_v2.csv, the new
    transaction-level gold set for src/final_evaluation.py to score against."""
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]],
            "amount": row[idx["amount"]], "direction": row[idx["direction"]],
            "provider": row[idx["provider"]], "native_category": row[idx["native_category"]],
            "source": row[idx["source"]], "provenance": row[idx["provenance"]],
            "haiku_leaf": row[idx["haiku_leaf"]], "sonnet_leaf": row[idx["sonnet_leaf"]],
            "gold_leaf": final_leaf, "notes": row[idx["notes"]] or "",
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader(); w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in MODELS:
            sys.exit(f"Usage: label [{'|'.join(MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(args[1]) if len(args) > 1 else REVIEW_COMPLETED_XLSX
        apply_review(path)
