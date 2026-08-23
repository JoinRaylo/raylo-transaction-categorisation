"""Gold set v5 -- the LOCKED test set (agreed with Carlos 2026-08-23).

## Why this exists

Every prior gold set has already been "spent" as a development/tuning set,
whether we meant it to be or not. `gold_v2_slm_eval_holdout.csv` was built as
a clean, merchant-disjoint holdout -- but it was then used to pick a winner
across FIVE prompt-compression variants (CLAUDE.md sec 6a) AND to pick the
production model (Gemini vs Opus vs Sonnet vs Haiku, sec 6a again) AND to
confirm the Option-1 consensus design. `gold_transactions_v4_slm_volume.csv`
was built to answer "does the model choice hold on real volume" and was
immediately used to confirm that same choice. None of this was wrong to do --
but it means every one of our eval sets has now been used to pick a winner at
least once, which is exactly the multiple-comparisons trap: the more ideas
you score against the same set and keep the best, the more that set's number
overstates how the winner will do on genuinely new data.

**This set is different by construction: it does not get scored against
anything during development.** No prompt variant, no model comparison, no
dictionary change gets checked against it. It exists for exactly one thing:
the actual go/no-go decision (Experiment 3's promotion call, or whatever the
next real milestone is) -- scored ONCE, at that decision point, then
considered spent itself and retired (build v6 for the next one).

## Population & sourcing

True random, whole-Plaid population (same methodology as
build_gold_v3_volume.py -- this is meant to validate the FULL pipeline
end-to-end at the actual decision point, not just the LLM/SLM tier, so it
should not be restricted to the unmatched residual the way v4 was).

Excludes every merchant that appears ANYWHERE in this project's existing
labelled data -- every prior gold set, the merchant dictionary, and every
production-labelling tranche -- so this is genuinely novel material nobody
has looked at, not just a fresh transaction-level sample of already-seen
merchants.

Two-model pair is Gemini 3.7 Flash + Sonnet 5 (Option 1, matching
production_labelling.py -- NOT Haiku, which the Option-1 refactor retired
from every live labelling path 2026-08-23).

## Usage

    python src/build_gold_v5_locked.py fetch
    python src/build_gold_v5_locked.py label gemini
    python src/build_gold_v5_locked.py label sonnet
    python src/build_gold_v5_locked.py sheet
    python src/build_gold_v5_locked.py apply [path]

Drafting (fetch/label/sheet) is fine to run now -- it doesn't touch the lock,
it just builds the ground truth. The lock is about SCORING: once
`data/gold_transactions_v5_LOCKED.csv` exists, do not run any benchmark
script against it until the actual decision point. If you're reading this
mid-project wondering whether it's OK to score against it "just to check" --
it isn't; that's exactly the discipline this file exists to protect.
"""
import csv
import json
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import (  # noqa: E402
    build_system_prompt, build_tool_schema, load_example_merchants, load_example_notes,
    build_notes_addendum, load_crosswalk,
)
from build_final_gold_v2 import TXN_ADDENDUM  # noqa: E402

SAMPLE_CSV = OUT_DIR / "gold_v5_locked_sample.csv"
# Option 1 pair (2026-08-23): Gemini 3.7 Flash + Sonnet 5, matching
# production_labelling.py.PRODUCTION_MODELS -- not gating_experiment.MODELS,
# which is the frozen historical Haiku-vs-Sonnet gating-experiment record.
V5_MODELS = {
    "gemini": {"backend": "gemini", "id": "gemini-3.7-flash", "extra": {}},
    "sonnet": {"backend": "anthropic", "id": "claude-sonnet-5", "max_tokens": 16000, "extra": {}},
}
PREDICTIONS = {k: OUT_DIR / f"gold_v5_locked_predictions_{k}.csv" for k in V5_MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v5_locked_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v5_locked_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v5_LOCKED.csv"

N_EQUIFAX = 400
N_PLAID = 700

EXISTING_GOLD_FILES = [
    "gold_transactions_v2.csv", "gold_transactions_v2_batch2.csv",
    "gold_transactions_v3_volume.csv", "gold_transactions_v4_slm_volume.csv",
    "gold_merchant_labels.csv", "gold_tail_labels.csv",
]
PRODUCTION_TRANCHE_FILES = [
    "production_labels_tranche1.csv", "production_labels_tranche2.csv", "production_labels_tranche3.csv",
]


def _norm(s):
    return (s or "").strip().lower()


def _seen_merchants():
    """Every merchant this project has ever looked at with a label attached --
    gold sets, production tranches, and the live dictionary. Excluded wholesale
    from v5 so this set is genuinely novel, not just fresh transactions of
    merchants we already have ground truth for."""
    seen = set()
    for fname in EXISTING_GOLD_FILES:
        path = ROOT / "data" / fname
        if path.exists():
            for r in csv.DictReader(open(path)):
                seen.add(_norm(r.get("merchant_raw") or r.get("merchant")))
    for fname in PRODUCTION_TRANCHE_FILES:
        path = ROOT / "data" / fname
        if path.exists():
            for r in csv.DictReader(open(path)):
                seen.add(_norm(r.get("merchant")))
    dict_path = ROOT / "taxonomy" / "merchant_dictionary.csv"
    if dict_path.exists():
        for r in csv.DictReader(open(dict_path)):
            seen.add(_norm(r.get("normalised_merchant")))
    seen.discard("")
    return seen


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    seen = _seen_merchants()
    print(f"{len(seen)} merchants already have a label somewhere in this project -- excluding all of them",
          file=sys.stderr)

    print(f"Sampling {N_EQUIFAX} Equifax + {N_PLAID} Plaid transactions, true random, novel merchants only...",
          file=sys.stderr)

    eqx_sql = f"""
    SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
           Description AS description_raw, Amount AS amount,
           IF(TransactionTypeId=1,'credit','debit') AS direction,
           CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category,
           'equifax' AS provider
    FROM `raylo-production.equifax_data.open_banking_full_dump`
    WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
    ORDER BY RAND()
    LIMIT {N_EQUIFAX * 40}
    """
    plaid_sql = f"""
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw, amount,
           IF(amount < 0, 'credit', 'debit') AS direction,
           credit_category_detailed AS native_category,
           'plaid' AS provider
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
    ORDER BY RAND()
    LIMIT {N_PLAID * 40}
    """

    all_rows = []
    for sql, target_n, label in [(eqx_sql, N_EQUIFAX, "equifax"), (plaid_sql, N_PLAID, "plaid")]:
        candidates = [dict(r) for r in client.query(sql).result()]
        novel = [r for r in candidates if _norm(r["merchant"]) not in seen]
        chosen = novel[:target_n]
        print(f"  [{label}] {len(candidates)} candidates -> {len(novel)} novel-merchant -> "
              f"{len(chosen)} chosen", file=sys.stderr)
        all_rows.extend(chosen)

    for i, r in enumerate(all_rows):
        r["row_id"] = i
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_rows)
    print(f"\nWrote {SAMPLE_CSV}: {len(all_rows)} rows, all genuinely novel merchants "
          f"(zero overlap with any existing gold set, production tranche, or the merchant dictionary)",
          file=sys.stderr)


def label(model_key):
    cfg = V5_MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TXN_ADDENDUM + build_notes_addendum(load_example_notes()))

    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)
    todo = [r for r in rows if r["row_id"] not in predictions]

    BATCH = 20

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow(p)

    if cfg["backend"] == "anthropic":
        import anthropic

        client = anthropic.Anthropic()
        tool = build_tool_schema(leaves)

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.messages.create(
                    model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                    system=system_prompt, tools=[tool],
                    tool_choice={"type": "tool", "name": "submit_classifications"},
                    messages=[{"role": "user", "content": user_msg}], timeout=90.0,
                    **cfg.get("extra", {}),
                )
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
            if not tool_use:
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in tool_use.input.get("results", []):
                r = by_idx.get(res.get("index"))
                if not r:
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"],
                                    "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
            return out

    elif cfg["backend"] == "gemini":
        import os

        from google import genai
        from google.genai import types

        client = genai.Client(api_key=os.environ["GOOGLE_API_KEY"], vertexai=False)
        leaf_list = sorted(leaves)
        index_addendum = "\n\n## Category index (output this number, not the name)\n" + "\n".join(
            f"{i + 1}. {leaf}" for i, leaf in enumerate(leaf_list))
        gemini_system = system_prompt + index_addendum
        schema = {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "index": {"type": "integer"},
                            "merchant": {"type": "string"},
                            "category_index": {"type": "integer", "minimum": 1, "maximum": len(leaf_list)},
                            "confidence": {"type": "number"},
                        },
                        "required": ["index", "merchant", "category_index", "confidence"],
                    },
                }
            },
            "required": ["results"],
        }

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.models.generate_content(
                    model=cfg["id"], contents=user_msg,
                    config=types.GenerateContentConfig(
                        system_instruction=gemini_system,
                        response_mime_type="application/json", response_schema=schema, temperature=0.0,
                    ),
                )
                data = json.loads(resp.text)
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in data.get("results", []):
                cat_idx = res.get("category_index")
                r = by_idx.get(res.get("index"))
                if not r or not (isinstance(cat_idx, int) and 1 <= cat_idx <= len(leaf_list)):
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"],
                                    "llm_leaf": leaf_list[cat_idx - 1], "llm_confidence": res.get("confidence")}
            return out

    else:
        sys.exit(f"unknown backend {cfg['backend']!r}")

    n_batches = (len(todo) + BATCH - 1) // BATCH
    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 5 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:03d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if r["row_id"] not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:03d}_r{attempt}"))
        if num % 10 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if r["row_id"] not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    gemini = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["gemini"]))} \
        if PREDICTIONS["gemini"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        k = r["row_id"]
        g, s = gemini.get(k), sonnet.get(k)
        r["gemini_leaf"] = g["llm_leaf"] if g else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        r["agree"] = "yes" if (g and s and g["llm_leaf"] == s["llm_leaf"]) else ("missing" if not (g and s) else "no")
        r["proposed_gold_leaf"] = g["llm_leaf"] if r["agree"] == "yes" else ""

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v5 -- THE LOCKED TEST SET -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A3"] = "What this is"
    ws_instr["B3"] = (
        f"{len(rows)} real transactions, true random, from merchants that have NEVER appeared in any prior "
        f"gold set, production tranche, or the merchant dictionary. Reviewing this builds the ground truth -- "
        f"that part is fine to do now, same as every other gold set.")
    ws_instr["A4"] = "The one rule that matters"
    ws_instr["B4"] = (
        "Once this file is built, it does NOT get scored against anything during ongoing development -- no "
        "prompt tweak, no model swap, no dictionary change gets checked against it. It gets scored exactly "
        "once, at the actual go/no-go decision (Experiment 3's promotion call or equivalent). Every other gold "
        "set we have has already been used to pick a winner at least once, which quietly overstates how well "
        "that winner generalises. This set exists specifically to not have that problem when it matters most.")
    ws_instr["A5"] = "What to do"
    ws_instr["B5"] = "Fill in `final_leaf` for every row (Taxonomy sheet has the valid list). Leave blank only if genuinely unclassifiable."
    ws_instr.column_dimensions["A"].width = 22
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "gemini_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf", "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["gemini_leaf"], r["sonnet_leaf"], r["agree"],
                   r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).fill = yellow
        ws.cell(row=row_idx, column=12).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 20, 20, 12, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows ({n_disagree} genuine disagreements)", file=sys.stderr)


def apply_review(path):
    """Produces data/gold_transactions_v5_LOCKED.csv, same schema as v2/v4's
    holdout files so it drops into existing scoring scripts -- but per the file
    name and every docstring in this module, DO NOT point a scoring script at
    it until the actual decision point."""
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        amount = row[idx["amount"]]
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]] or "",
            "amount": abs(float(amount)), "direction": row[idx["direction"]],
            "gold_leaf": final_leaf,
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant_raw", "description_raw", "amount", "direction", "gold_leaf"])
        w.writeheader()
        w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)
    print("REMINDER: this is the locked test set. Do not score anything against it "
          "until the actual go/no-go decision.", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in V5_MODELS:
            sys.exit(f"Usage: label [{'|'.join(V5_MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
