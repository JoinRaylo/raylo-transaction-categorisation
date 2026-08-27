"""Gold set v6 -- the LOCKED confirmation set (replaces v5, 2026-08-26).

v5 was novel at construction, then tranche 4 and the 91k dictionary covered
hundreds of those merchants. v5 stays in git as reviewed labels and is not
scored. This set is the replacement: same methodology (400 Equifax + 700
Plaid, true random), same scoring rule (once, at go/no-go), with a complete
exclusion list — every prior gold set including v5, risk gold, holdout,
eyeball CSVs, leaf top-up, production tranches 1–4, and the live dictionary.

Do not score `data/gold_transactions_v6_LOCKED.csv` during development.

## Usage

    python src/build_gold_v6_locked.py fetch
    python src/build_gold_v6_locked.py label gemini
    python src/build_gold_v6_locked.py label sonnet
    python src/build_gold_v6_locked.py sheet
    python src/build_gold_v6_locked.py apply [path]

Fetch writes `outputs/gold_v6_locked_sample.csv` (gitignored). Labelling is
Gemini 3.7 Flash + Sonnet 5 (Option 1). Do not fire 1,100 LLM calls unless
asked — fetch first so membership is frozen before the dictionary grows again.
"""
import csv
import json
import pathlib
import sys

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import (  # noqa: E402
    build_system_prompt, build_tool_schema, load_example_merchants, load_example_notes,
    build_notes_addendum, load_crosswalk,
)
from build_final_gold_v2 import TXN_ADDENDUM  # noqa: E402
from eval_sets import v6_excluded_merchants  # noqa: E402

SAMPLE_CSV = OUT_DIR / "gold_v6_locked_sample.csv"
V6_MODELS = {
    "gemini": {"backend": "gemini", "id": "gemini-3.7-flash", "extra": {}},
    "sonnet": {"backend": "anthropic", "id": "claude-sonnet-5", "max_tokens": 16000, "extra": {}},
}
PREDICTIONS = {k: OUT_DIR / f"gold_v6_locked_predictions_{k}.csv" for k in V6_MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v6_locked_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v6_locked_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v6_LOCKED.csv"

N_EQUIFAX = 400
N_PLAID = 700
# Dictionary is ~91k keys; oversample harder than v5 so Equifax still yields
# novel vendors after those exclusions.
OVERSAMPLE = 200


def _norm(s):
    return (s or "").strip().lower()


def _seen_merchants():
    return v6_excluded_merchants()


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    seen = _seen_merchants()
    print(f"{len(seen)} merchants already have a label somewhere in this project -- excluding all of them",
          file=sys.stderr)

    print(f"Sampling {N_EQUIFAX} Equifax + {N_PLAID} Plaid transactions, true random, novel merchants only...",
          file=sys.stderr)

    eqx_sql = f"""
    SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
           Description AS description_raw, Amount AS amount,
           IF(TransactionTypeId=1,'credit','debit') AS direction,
           CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category,
           'equifax' AS provider
    FROM `raylo-production.equifax_data.open_banking_full_dump`
    WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
    ORDER BY RAND()
    LIMIT {N_EQUIFAX * OVERSAMPLE}
    """
    plaid_sql = f"""
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw, amount,
           IF(amount < 0, 'credit', 'debit') AS direction,
           credit_category_detailed AS native_category,
           'plaid' AS provider
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
    ORDER BY RAND()
    LIMIT {N_PLAID * OVERSAMPLE}
    """

    all_rows = []
    for sql, target_n, label in [(eqx_sql, N_EQUIFAX, "equifax"), (plaid_sql, N_PLAID, "plaid")]:
        candidates = [dict(r) for r in client.query(sql).result()]
        novel = [r for r in candidates if _norm(r["merchant"]) not in seen]
        chosen, seen_in_batch = [], set()
        for r in novel:
            m = _norm(r["merchant"])
            if m in seen_in_batch:
                continue
            seen_in_batch.add(m)
            chosen.append(r)
            if len(chosen) >= target_n:
                break
        seen |= seen_in_batch
        if len(chosen) < target_n:
            print(f"  WARN [{label}] only {len(chosen)}/{target_n} novel merchants",
                  file=sys.stderr)
        print(f"  [{label}] {len(candidates)} candidates -> {len(novel)} novel-merchant -> "
              f"{len(chosen)} chosen", file=sys.stderr)
        all_rows.extend(chosen)

    for i, r in enumerate(all_rows):
        r["row_id"] = i
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_rows)
    print(f"\nWrote {SAMPLE_CSV}: {len(all_rows)} rows, all genuinely novel merchants "
          f"(zero overlap with any existing gold set, production tranche, or the merchant dictionary)",
          file=sys.stderr)


def label(model_key):
    cfg = V6_MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TXN_ADDENDUM + build_notes_addendum(load_example_notes()))

    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)
    todo = [r for r in rows if r["row_id"] not in predictions]

    BATCH = 20

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow(p)

    if cfg["backend"] == "anthropic":
        import anthropic

        client = anthropic.Anthropic()
        tool = build_tool_schema(leaves)

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.messages.create(
                    model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                    system=system_prompt, tools=[tool],
                    tool_choice={"type": "tool", "name": "submit_classifications"},
                    messages=[{"role": "user", "content": user_msg}], timeout=90.0,
                    **cfg.get("extra", {}),
                )
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
            if not tool_use:
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in tool_use.input.get("results", []):
                r = by_idx.get(res.get("index"))
                if not r:
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"],
                                    "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
            return out

    elif cfg["backend"] == "gemini":
        import os

        from google import genai
        from google.genai import types

        client = genai.Client(api_key=os.environ["GOOGLE_API_KEY"], vertexai=False)
        leaf_list = sorted(leaves)
        index_addendum = "\n\n## Category index (output this number, not the name)\n" + "\n".join(
            f"{i + 1}. {leaf}" for i, leaf in enumerate(leaf_list))
        gemini_system = system_prompt + index_addendum
        schema = {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "index": {"type": "integer"},
                            "merchant": {"type": "string"},
                            "category_index": {"type": "integer", "minimum": 1, "maximum": len(leaf_list)},
                            "confidence": {"type": "number"},
                        },
                        "required": ["index", "merchant", "category_index", "confidence"],
                    },
                }
            },
            "required": ["results"],
        }

        def classify_batch(batch, tag, attempt=0):
            user_msg = ("Classify each of these real transactions:\n\n"
                        + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
            try:
                resp = client.models.generate_content(
                    model=cfg["id"], contents=user_msg,
                    config=types.GenerateContentConfig(
                        system_instruction=gemini_system,
                        response_mime_type="application/json", response_schema=schema, temperature=0.0,
                    ),
                )
                data = json.loads(resp.text)
            except Exception as e:
                if attempt < 2:
                    print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                    import time
                    time.sleep(2 ** attempt)
                    return classify_batch(batch, tag, attempt + 1)
                print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
                return {}
            by_idx = {j + 1: r for j, r in enumerate(batch)}
            out = {}
            for res in data.get("results", []):
                cat_idx = res.get("category_index")
                r = by_idx.get(res.get("index"))
                if not r or not (isinstance(cat_idx, int) and 1 <= cat_idx <= len(leaf_list)):
                    continue
                out[r["row_id"]] = {"row_id": r["row_id"], "merchant": r["merchant"],
                                    "llm_leaf": leaf_list[cat_idx - 1], "llm_confidence": res.get("confidence")}
            return out

    else:
        sys.exit(f"unknown backend {cfg['backend']!r}")

    n_batches = (len(todo) + BATCH - 1) // BATCH
    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 5 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:03d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if r["row_id"] not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:03d}_r{attempt}"))
        if num % 10 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if r["row_id"] not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    gemini = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["gemini"]))} \
        if PREDICTIONS["gemini"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        k = r["row_id"]
        g, s = gemini.get(k), sonnet.get(k)
        r["gemini_leaf"] = g["llm_leaf"] if g else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        r["agree"] = "yes" if (g and s and g["llm_leaf"] == s["llm_leaf"]) else ("missing" if not (g and s) else "no")
        r["proposed_gold_leaf"] = g["llm_leaf"] if r["agree"] == "yes" else ""

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v6 -- THE LOCKED TEST SET -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A3"] = "What this is"
    ws_instr["B3"] = (
        f"{len(rows)} real transactions, true random, from merchants that have NEVER appeared in v5, "
        f"any other gold set, production tranche 1–4, or the merchant dictionary. Reviewing this "
        f"builds the ground truth -- that part is fine to do now.")
    ws_instr["A4"] = "The one rule that matters"
    ws_instr["B4"] = (
        "Once this file is built, it does NOT get scored against anything during ongoing development -- no "
        "prompt tweak, no model swap, no dictionary change gets checked against it. It gets scored exactly "
        "once, at the actual go/no-go decision (Experiment 3's promotion call or equivalent). Every other gold "
        "set we have has already been used to pick a winner at least once, which quietly overstates how well "
        "that winner generalises. This set exists specifically to not have that problem when it matters most.")
    ws_instr["A5"] = "What to do"
    ws_instr["B5"] = "Fill in `final_leaf` for every row (Taxonomy sheet has the valid list). Leave blank only if genuinely unclassifiable."
    ws_instr.column_dimensions["A"].width = 22
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "gemini_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf", "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["gemini_leaf"], r["sonnet_leaf"], r["agree"],
                   r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).fill = yellow
        ws.cell(row=row_idx, column=12).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 20, 20, 12, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows ({n_disagree} genuine disagreements)", file=sys.stderr)


def apply_review(path):
    """Produces data/gold_transactions_v6_LOCKED.csv. Same schema as v2/v4
    holdout files. Do not point a scoring script at it until go/no-go."""
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    empty_merchant_kept = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        description_raw = row[idx["description_raw"]]
        # End-of-sheet rows have BOTH fields empty. A blank merchant_raw alone is
        # real source data (Plaid merchant_name is only 63.4% filled) and the row
        # can still carry a genuinely reviewed label off the narrative alone --
        # do not drop it just because merchant_raw is blank (see
        # build_gold_risk_categories.py, where this exact bug dropped 84/711
        # reviewed rows on the first run).
        if merchant_raw is None and not description_raw:
            continue
        if merchant_raw is None:
            merchant_raw = ""
            empty_merchant_kept += 1
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        amount = row[idx["amount"]]
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": description_raw or "",
            "amount": abs(float(amount)), "direction": row[idx["direction"]],
            "gold_leaf": final_leaf,
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["merchant_raw", "description_raw", "amount", "direction", "gold_leaf"])
        w.writeheader()
        w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded; {empty_merchant_kept} kept with no merchant field, "
          f"reviewed off the narrative alone)", file=sys.stderr)
    print("REMINDER: this is the locked test set. Do not score anything against it "
          "until the actual go/no-go decision.", file=sys.stderr)


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in V6_MODELS:
            sys.exit(f"Usage: label [{'|'.join(V6_MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
