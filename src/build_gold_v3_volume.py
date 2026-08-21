"""Gold set v3 -- volume-weighted evaluation.

v2 (build_final_gold_v2*.py) deliberately stratified by native provider
category, capping each category so rare ones weren't crowded out. That's
right for measuring breadth across the taxonomy, but it means a category
with 10M real transactions and one with 500 got roughly equal sample
weight -- nothing like production traffic, which is extremely
concentrated (top 1k Plaid merchant strings = 44.9% of unmatched volume).

v3 answers a different, complementary question: **what fraction of ACTUAL
transaction volume gets classified correctly today?** Sampling is genuinely
random over the raw population (no stratification, no per-category cap),
so high-volume merchants dominate the sample in roughly their true
proportion, same as live traffic would.

To keep the review burden proportional to what's actually NEW: any sampled
transaction whose merchant already has an established, non-conflicting
gold_leaf from v2 (batch 1 or 2) is auto-resolved without asking Carlos to
re-review it -- only genuinely new merchants go into the review workbook.
Conflicting merchants (revolut, monzo, ... -- proven context-dependent by
v2's own data) are never auto-resolved; they always need a fresh per-
transaction look, same as any other new row.

Equifax and Plaid are sampled and reported SEPARATELY, not blended --
mixing them would require an arbitrary provider-weighting choice, and
Plaid is what matters for live production traffic while Equifax is dead
history, so conflating them would obscure the number that actually matters.

Usage:
    python src/build_gold_v3_volume.py fetch
    python src/build_gold_v3_volume.py label haiku
    python src/build_gold_v3_volume.py label sonnet
    python src/build_gold_v3_volume.py sheet
    python src/build_gold_v3_volume.py apply [path]
    python src/build_gold_v3_volume.py score   # after apply
"""
import csv
import pathlib
import sys
from collections import defaultdict

from dotenv import load_dotenv

load_dotenv()

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "outputs"
sys.path.insert(0, str(ROOT / "src"))
from gating_experiment import (MODELS, build_system_prompt, build_tool_schema,  # noqa: E402
                                load_example_merchants, load_example_notes, build_notes_addendum, load_crosswalk)
from build_final_gold_v2 import TXN_ADDENDUM  # noqa: E402
import final_evaluation as fe  # noqa: E402
fe.SUB_MAP, fe.PRI_MAP, fe.PLAID_MAP, _ = fe.load_crosswalk()
fe.DICTIONARY = fe.load_dictionary()
fe.RULES = fe.load_rules()
eqx_native_leaf, plaid_native_leaf, our_leaf = fe.eqx_native_leaf, fe.plaid_native_leaf, fe.our_leaf

SAMPLE_CSV = OUT_DIR / "gold_v3_sample.csv"
PREDICTIONS = {k: OUT_DIR / f"gold_v3_predictions_{k}.csv" for k in MODELS}
REVIEW_XLSX = OUT_DIR / "gold_v3_review.xlsx"
REVIEW_COMPLETED_XLSX = OUT_DIR / "gold_v3_review_completed.xlsx"
FINAL_CSV = ROOT / "data" / "gold_transactions_v3_volume.csv"
V3_REPORT_MD = ROOT / "data" / "final_evaluation_v3_volume_report.md"

N_EQUIFAX = 600
N_PLAID = 900


def _norm(s):
    return (s or "").strip().lower()


def _established_truth():
    """merchant -> gold_leaf for every v2 merchant with a SINGLE consistent leaf across
    both batches. Merchants with conflicting leaves (proven context-dependent) are
    deliberately excluded -- a new transaction for them still needs its own review."""
    by_merchant = defaultdict(set)
    for f in ["gold_transactions_v2.csv", "gold_transactions_v2_batch2.csv"]:
        path = ROOT / "data" / f
        if path.exists():
            for r in csv.DictReader(open(path)):
                by_merchant[_norm(r["merchant_raw"])].add(r["gold_leaf"])
    return {m: next(iter(leaves)) for m, leaves in by_merchant.items() if len(leaves) == 1}


def fetch():
    from google.cloud import bigquery
    client = bigquery.Client(project="raylo-production")

    established = _established_truth()
    print(f"{len(established)} merchants have an established, non-conflicting v2 gold_leaf "
          f"-- will be auto-resolved if sampled", file=sys.stderr)

    print(f"Sampling {N_EQUIFAX} Equifax transactions, true random (no stratification)...", file=sys.stderr)
    eqx_sql = f"""
    SELECT LOWER(TRIM(VendorDescription)) AS merchant, VendorDescription AS merchant_raw,
           Description AS description_raw, Amount AS amount,
           IF(TransactionTypeId=1,'credit','debit') AS direction,
           CONCAT(PrimaryCategoryDescription, ' | ', SubCategoryDescription) AS native_category
    FROM `raylo-production.equifax_data.open_banking_full_dump`
    TABLESAMPLE SYSTEM (1 PERCENT)
    WHERE VendorDescription IS NOT NULL AND TRIM(VendorDescription) != ''
    ORDER BY RAND()
    LIMIT {N_EQUIFAX}
    """
    eqx_rows = [dict(r) for r in client.query(eqx_sql).result()]

    print(f"Sampling {N_PLAID} Plaid transactions, true random (no stratification)...", file=sys.stderr)
    plaid_sql = f"""
    SELECT LOWER(TRIM(merchant_name)) AS merchant, merchant_name AS merchant_raw,
           COALESCE(original_description, transaction_name) AS description_raw,
           amount, IF(amount < 0,'credit','debit') AS direction,
           credit_category_detailed AS native_category
    FROM `raylo-production.dbt_production.credit_plaid_open_banking_transactions`
    WHERE merchant_name IS NOT NULL AND TRIM(merchant_name) != ''
    ORDER BY RAND()
    LIMIT {N_PLAID}
    """
    plaid_rows = [dict(r) for r in client.query(plaid_sql).result()]

    all_rows = []
    for provider, rows in [("equifax", eqx_rows), ("plaid", plaid_rows)]:
        for r in rows:
            m = r["merchant"]
            all_rows.append({
                "merchant": m, "merchant_raw": r["merchant_raw"],
                "description_raw": r["description_raw"] or "", "amount": r["amount"],
                "direction": r["direction"], "native_category": r["native_category"],
                "provider": provider,
                "established_leaf": established.get(m, ""),
                "haiku_leaf": "", "sonnet_leaf": "", "agree": "", "proposed_gold_leaf": "",
            })

    for i, r in enumerate(all_rows):
        r["row_id"] = i
    n_established = sum(1 for r in all_rows if r["established_leaf"])
    OUT_DIR.mkdir(exist_ok=True)
    fieldnames = ["row_id"] + [k for k in all_rows[0].keys() if k != "row_id"]
    with open(SAMPLE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(all_rows)
    print(f"Wrote {SAMPLE_CSV}: {len(all_rows)} rows ({len(eqx_rows)} eqx / {len(plaid_rows)} plaid), "
          f"{n_established} already resolvable from established v2 truth, "
          f"{len(all_rows) - n_established} need fresh labelling", file=sys.stderr)


def label(model_key):
    import anthropic

    cfg = MODELS[model_key]
    _, _, leaves, gen_of, notes_of = load_crosswalk()
    system_prompt = (build_system_prompt(leaves, gen_of, notes_of, load_example_merchants())
                      + TXN_ADDENDUM + build_notes_addendum(load_example_notes()))
    tool = build_tool_schema(leaves)

    rows = [r for r in csv.DictReader(open(SAMPLE_CSV)) if not r["established_leaf"]]
    out_path = PREDICTIONS[model_key]
    predictions = {}
    if out_path.exists():
        predictions = {r["row_id"]: r for r in csv.DictReader(open(out_path)) if r["llm_leaf"]}
        print(f"Resuming: {len(predictions)} already labelled", file=sys.stderr)

    def row_key(r):
        return r["row_id"]

    todo = [r for r in rows if row_key(r) not in predictions]
    client = anthropic.Anthropic()
    BATCH = 20
    n_batches = (len(todo) + BATCH - 1) // BATCH

    def render(i, r):
        return (f"{i}. merchant: {r['merchant_raw']}\n"
                f"   description: {r['description_raw']}\n"
                f"   amount_gbp: {r['amount']} | direction: {r['direction']}\n"
                f"   native_category: {r['native_category']}")

    def flush():
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row_id", "merchant", "amount", "llm_leaf", "llm_confidence"])
            w.writeheader()
            for k, p in predictions.items():
                w.writerow({"row_id": p["row_id"], "merchant": p["merchant"], "amount": p["amount"],
                            "llm_leaf": p["llm_leaf"], "llm_confidence": p["llm_confidence"]})

    def classify_batch(batch, tag, attempt=0):
        user_msg = ("Classify each of these real transactions:\n\n"
                    + "\n".join(render(j + 1, r) for j, r in enumerate(batch)))
        try:
            resp = client.messages.create(
                model=cfg["id"], max_tokens=cfg.get("max_tokens", 8000),
                system=system_prompt, tools=[tool],
                tool_choice={"type": "tool", "name": "submit_classifications"},
                messages=[{"role": "user", "content": user_msg}],
                timeout=90.0,
                **cfg.get("extra", {}),
            )
        except Exception as e:
            if attempt < 2:
                print(f"  [{tag}] error ({e}), retrying...", file=sys.stderr)
                import time; time.sleep(2 ** attempt)
                return classify_batch(batch, tag, attempt + 1)
            print(f"  [{tag}] FAILED after retries: {e}", file=sys.stderr)
            return {}
        tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
        if not tool_use:
            return {}
        by_idx = {j + 1: r for j, r in enumerate(batch)}
        out = {}
        for res in tool_use.input.get("results", []):
            idx = res.get("index")
            r = by_idx.get(idx)
            if not r:
                continue
            out[row_key(r)] = {"row_id": r["row_id"], "merchant": r["merchant"], "amount": r["amount"],
                                "llm_leaf": res.get("detailed_category"), "llm_confidence": res.get("confidence")}
        return out

    for i in range(0, len(todo), BATCH):
        batch = todo[i:i + BATCH]
        num = i // BATCH + 1
        if num % 10 == 1:
            print(f"[{model_key}] batch {num}/{n_batches}", file=sys.stderr)
        predictions.update(classify_batch(batch, f"b{num:04d}"))
        for attempt in (1, 2):
            missing = [r for r in batch if row_key(r) not in predictions]
            if not missing:
                break
            predictions.update(classify_batch(missing, f"b{num:04d}_r{attempt}"))
        if num % 15 == 0:
            flush()
    flush()
    missing = sum(1 for r in todo if row_key(r) not in predictions)
    print(f"Wrote {out_path}: {len(predictions)} labelled, {missing} missing", file=sys.stderr)


def sheet():
    rows = list(csv.DictReader(open(SAMPLE_CSV)))
    haiku = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["haiku"]))} \
        if PREDICTIONS["haiku"].exists() else {}
    sonnet = {r["row_id"]: r for r in csv.DictReader(open(PREDICTIONS["sonnet"]))} \
        if PREDICTIONS["sonnet"].exists() else {}

    for r in rows:
        if r["established_leaf"]:
            r["proposed_gold_leaf"] = r["established_leaf"]
            r["agree"] = "established"
            continue
        k = r["row_id"]
        h, s = haiku.get(k), sonnet.get(k)
        r["haiku_leaf"] = h["llm_leaf"] if h else ""
        r["sonnet_leaf"] = s["llm_leaf"] if s else ""
        if h and s and h["llm_leaf"] == s["llm_leaf"]:
            r["proposed_gold_leaf"] = h["llm_leaf"]
            r["agree"] = "yes"
        elif h and s:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "no"
        else:
            r["proposed_gold_leaf"] = ""
            r["agree"] = "missing"

    _, _, leaves, gen_of, _ = load_crosswalk()

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Gold transaction set v3 (volume-weighted) -- review instructions"
    ws_instr["A1"].font = Font(bold=True, size=13)
    n_established = sum(1 for r in rows if r["agree"] == "established")
    n_review = len(rows) - n_established
    ws_instr["A3"] = "What this is"
    ws_instr["B3"] = (f"A TRUE random sample of real transactions (no category stratification this time), "
                      f"so common merchants dominate roughly like real traffic. {n_established} of {len(rows)} "
                      f"rows already have an established, non-conflicting answer from the earlier v2 review -- "
                      f"marked agree='established', final_leaf pre-filled, spot-check only. {n_review} rows are "
                      f"genuinely new merchants needing your review, same process as before.")
    ws_instr["A4"] = "What to do"
    ws_instr["B4"] = ("Fill in `final_leaf` for every row (Taxonomy sheet has the valid list). Leave blank only "
                      "if genuinely unclassifiable.")
    ws_instr.column_dimensions["A"].width = 18
    ws_instr.column_dimensions["B"].width = 100

    ws = wb.create_sheet("Review")
    header = ["merchant_raw", "description_raw", "amount", "direction", "provider", "native_category",
              "haiku_leaf", "sonnet_leaf", "agree", "proposed_gold_leaf", "final_leaf", "notes"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["merchant_raw"], r["description_raw"], r["amount"], r["direction"], r["provider"],
                   r["native_category"], r["haiku_leaf"], r["sonnet_leaf"], r["agree"],
                   r["proposed_gold_leaf"], r["proposed_gold_leaf"], ""])
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).fill = yellow
        ws.cell(row=row_idx, column=12).fill = yellow
    dv = DataValidation(type="list", formula1=f'"{",".join(leaves)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    widths = [22, 45, 10, 9, 8, 28, 20, 20, 12, 22, 22, 30]
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w

    ws_tax = wb.create_sheet("Taxonomy")
    ws_tax.append(["detailed_category", "general_category"])
    for leaf in leaves:
        ws_tax.append([leaf, gen_of[leaf]])

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(REVIEW_XLSX)
    n_disagree = sum(1 for r in rows if r["agree"] == "no")
    print(f"Wrote {REVIEW_XLSX}: {len(rows)} rows ({n_established} pre-resolved, "
          f"{n_disagree} genuine disagreements)", file=sys.stderr)


def apply_review(path):
    import openpyxl

    _, _, leaves, gen_of, _ = load_crosswalk()
    ws = openpyxl.load_workbook(path, data_only=True)["Review"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_rows = []
    blank = 0
    bad_leaf = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_raw = row[idx["merchant_raw"]]
        if merchant_raw is None:
            continue
        final_leaf = row[idx["final_leaf"]]
        if not final_leaf:
            blank += 1
            continue
        if final_leaf not in gen_of:
            bad_leaf.append((merchant_raw, final_leaf))
            continue
        out_rows.append({
            "merchant_raw": merchant_raw, "description_raw": row[idx["description_raw"]],
            "amount": row[idx["amount"]], "direction": row[idx["direction"]],
            "provider": row[idx["provider"]], "native_category": row[idx["native_category"]],
            "agree": row[idx["agree"]], "gold_leaf": final_leaf, "notes": row[idx["notes"]] or "",
        })

    if bad_leaf:
        sys.exit(f"{len(bad_leaf)} rows have a final_leaf not in the taxonomy: {bad_leaf[:10]}")

    FINAL_CSV.parent.mkdir(exist_ok=True)
    with open(FINAL_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader(); w.writerows(out_rows)
    print(f"Wrote {FINAL_CSV}: {len(out_rows)} gold transactions "
          f"({blank} left blank/unclassifiable, excluded)", file=sys.stderr)


def score():
    _, _, _, gen_of = load_crosswalk()
    self_sourced = {r["normalised_merchant"] for r in csv.DictReader(open(ROOT / "taxonomy" / "merchant_dictionary.csv"))
                    if r["source"] == "gold_v2_review"}

    rows = list(csv.DictReader(open(FINAL_CSV)))
    out_rows = []
    for r in rows:
        provider, direction = r["provider"], r["direction"]
        merchant, description, gold_leaf = r["merchant_raw"], r["description_raw"], r["gold_leaf"]
        if provider == "equifax":
            pri, sub = (r["native_category"].split(" | ", 1) + [""])[:2] if r["native_category"] else ("", "")
            native_leaf = eqx_native_leaf(pri, sub, direction)
            our, tier = our_leaf(merchant, direction, description, eqx_native_leaf, pri, sub, direction)
        else:
            cat = r["native_category"]
            native_leaf = plaid_native_leaf(cat, direction)
            our, tier = our_leaf(merchant, direction, description, plaid_native_leaf, cat, direction)
        out_rows.append({"merchant_raw": merchant, "provider": provider, "gold_leaf": gold_leaf,
                          "native_leaf": native_leaf, "our_leaf": our, "our_tier": tier,
                          "self_sourced_dict_entry": merchant.strip().lower() in self_sourced})

    scoring_rows = [r for r in out_rows if not r["self_sourced_dict_entry"]]

    def acc(subset, key, level="leaf"):
        scored = [r for r in subset if r[key]]
        if not scored:
            return None, 0
        if level == "leaf":
            correct = sum(1 for r in scored if r[key] == r["gold_leaf"])
        else:
            correct = sum(1 for r in scored if gen_of.get(r[key]) == gen_of.get(r["gold_leaf"]))
        return correct / len(scored), len(scored)

    lines = ["# Final evaluation v3: volume-weighted (true random sample, no category stratification)\n",
             f"{len(out_rows)} real transactions, true random sample (Equifax + Plaid sampled and reported "
             f"separately -- see below). {len(out_rows) - len(scoring_rows)} excluded from scoring as "
             f"self-sourced dictionary entries. This answers a different question from the v2 breadth "
             f"evaluation: **what fraction of actual transaction VOLUME gets classified correctly today**, "
             f"since high-volume merchants dominate this sample in roughly their true proportion.\n"]
    for label_name, provider in [("Equifax", "equifax"), ("Plaid", "plaid")]:
        subset = [r for r in scoring_rows if r["provider"] == provider]
        if not subset:
            continue
        na, nn = acc(subset, "native_leaf", "leaf")
        ng, _ = acc(subset, "native_leaf", "general")
        oa, on = acc(subset, "our_leaf", "leaf")
        og, _ = acc(subset, "our_leaf", "general")
        lines.append(f"## {label_name} (n={len(subset)})\n")
        lines.append("| Source | Leaf accuracy | General-category accuracy |")
        lines.append("|---|---|---|")
        lines.append(f"| Native provider category | {na:.1%} | {ng:.1%} |")
        lines.append(f"| Our pipeline | {oa:.1%} | {og:.1%} |\n")

    V3_REPORT_MD.write_text("\n".join(lines) + "\n")
    print(f"Wrote {V3_REPORT_MD}", file=sys.stderr)
    print("\n".join(lines))


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] not in {"fetch", "label", "sheet", "apply", "score"}:
        sys.exit(__doc__)
    if args[0] == "fetch":
        fetch()
    elif args[0] == "label":
        if len(sys.argv) < 3 or sys.argv[2] not in MODELS:
            sys.exit(f"Usage: label [{'|'.join(MODELS)}]")
        label(sys.argv[2])
    elif args[0] == "sheet":
        sheet()
    elif args[0] == "apply":
        path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else REVIEW_COMPLETED_XLSX
        apply_review(path)
    elif args[0] == "score":
        score()
